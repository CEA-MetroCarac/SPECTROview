"""
This module contains all common methods or Class which are shared between 3 modules (Map, Spectra, Visualization).
"""
import markdown
import platform
import os
import shutil
import tempfile
import re
import json
import base64
import zlib
import pandas as pd
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
import matplotlib.patches as patches
import matplotlib.colors as mcolors
import matplotlib.cm as cm

from matplotlib.figure import Figure
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT

from io import BytesIO
from PIL import Image
from copy import deepcopy
from threading import Thread
from openpyxl.styles import PatternFill
from scipy.interpolate import RegularGridInterpolator
from scipy.interpolate import griddata
from superqt import QLabeledDoubleRangeSlider 

from app import PEAK_MODELS, PALETTE, DEFAULT_COLORS, DEFAULT_MARKERS, MARKERS, X_AXIS_UNIT, ICON_DIR, PLOT_POLICY
from fitspy.core.utils_mp import fit_mp
from fitspy.core.spectrum import Spectrum
from fitspy.core.spectra import Spectra
from fitspy.core.baseline import BaseLine

from multiprocessing import Queue

if platform.system() == 'Darwin':
    import AppKit 
if platform.system() == 'Windows':
    import win32clipboard

from PySide6.QtWidgets import QMessageBox, QDialog, QTableWidget,QWidgetAction, \
    QTableWidgetItem, QVBoxLayout, QHBoxLayout, QTextBrowser, QLabel, QToolButton, \
    QLineEdit, QWidget, QPushButton, QComboBox, QCheckBox, QListWidgetItem, QFileDialog,\
    QApplication,  QWidget, QMenu, QStyledItemDelegate, QListWidget, QAbstractItemView, QSizePolicy, QGroupBox, QFrame, QSpacerItem, QStyledItemDelegate
from PySide6.QtCore import Signal, QThread, Qt, QSize, QCoreApplication, QFileInfo,QPoint
from PySide6.QtGui import QPalette, QColor, QTextCursor, QIcon, QAction, Qt, QPixmap, QImage,QCursor

# Define a dictionary mapping RGBA tuples to named colors
rgba_to_named_color_dict = {mcolors.to_rgba(color_name): color_name for
                            color_name in mcolors.CSS4_COLORS}


class MapViewWidget(QWidget):
    """Class to manage the 2Dmap view widget"""

    def __init__(self, main_app, app_settings):
        super().__init__()
        self.main_app = main_app
        self.app_settings = app_settings  

        self.map_df_name = None
        self.map_df =pd.DataFrame() 
        self.df_fit_results =pd.DataFrame() 
        self.map_type = '2Dmap'
        # self.map_type = getattr(self.app_settings, "map_type", None)

        self.dpi = 70
        self.figure = None
        self.ax = None
        self.canvas = None
        self.toolbar = None
        self.spectra_listbox = None
        self.menu_actions = {}
        self.initUI()

    def initUI(self):
        """Initialize the UI components."""
        self.widget = QWidget()
        self.widget.setFixedWidth(400)
        self.map_widget_layout = QVBoxLayout(self.widget)
        self.map_widget_layout.setContentsMargins(0, 0, 0, 0) 
        self.create_widget()
    

    def create_widget(self):
        """Create 2Dmap plot widgets"""
        # Create a frame to hold the canvas with a fixed size
        self.canvas_frame = QFrame(self.widget)
        self.canvas_frame.setFixedSize(400, 350)
        frame_layout = QVBoxLayout(self.canvas_frame)
        frame_layout.setContentsMargins(5, 0, 5, 0)

        self.figure = plt.figure(dpi=70)
        self.ax = self.figure.add_subplot(111)
        
        self.ax.tick_params(axis='x', which='both')
        self.ax.tick_params(axis='y', which='both')
        self.canvas = FigureCanvas(self.figure)
        self.toolbar =NavigationToolbar2QT(self.canvas)
        for action in self.toolbar.actions():
            if action.text() in ['Customize','Zoom','Save', 'Pan', 'Back', 'Forward', 'Subplots']:
                action.setVisible(False)

        frame_layout.addWidget(self.canvas)
        toolbar_layout=QHBoxLayout()
        toolbar_layout.setContentsMargins(5, 0, 5, 0)
        
        spacer = QSpacerItem(40, 20, QSizePolicy.Expanding, QSizePolicy.Minimum)
        self.btn_copy = QPushButton("", self)
        icon = QIcon()
        icon.addFile(os.path.join(ICON_DIR, "copy.png"))
        self.btn_copy.setIcon(icon)
        self.btn_copy.setIconSize(QSize(24, 24))
        self.btn_copy.clicked.connect(self.copy_fig)
        
        
        
        toolbar_layout.addWidget(self.toolbar)
        toolbar_layout.addItem(spacer)
        
        toolbar_layout.addWidget(self.btn_copy)

        frame_layout.addLayout(toolbar_layout)

        # Add the map frame to the main layout
        self.map_widget_layout.addWidget(self.canvas_frame)

        # Variables to keep track of highlighted points and Ctrl key status
        self.selected_points = []
        self.ctrl_pressed = False
        
        # Connect the mouse and key events to the handler functions
        self.figure.canvas.mpl_connect('button_press_event', self.on_left_click_2Dmap)
        self.figure.canvas.mpl_connect('key_press_event', self.on_key_press)
        self.figure.canvas.mpl_connect('key_release_event', self.on_key_release)
        
        self.canvas.draw_idle()

        #### MAP_TYPE ComboBox (wafer or 2Dmap)
        combobox_layout = QHBoxLayout()
        self.map_type_label = QLabel("Map Type:")
        self.cbb_map_type = QComboBox(self)
        self.cbb_map_type.setFixedWidth(93)
        self.cbb_map_type.addItems(['2Dmap', 'Wafer_300mm', 'Wafer_200mm', 'Wafer_100mm'])
        
        
        self.cbb_map_type.currentIndexChanged.connect(self.refresh_plot)

        # Load last saved settings
        self.cbb_map_type.currentIndexChanged.connect(self.update_settings)
        saved_map_type = self.app_settings.map_type
        
        # Set the saved values in the combo boxes
        if saved_map_type in ['2Dmap', 'Wafer_300mm', 'Wafer_200mm', 'Wafer_100mm']:
            self.cbb_map_type.setCurrentText(saved_map_type)
        
        # Palette selector with preview 
        self.cbb_palette = CustomizedPalette()
        self.cbb_palette.currentIndexChanged.connect(self.refresh_plot)

        self.cb_auto_scale = QCheckBox("Auto Scale")
        self.cb_auto_scale.setChecked(True)
        self.cb_auto_scale.stateChanged.connect(self.update_z_range_slider)
        self.cb_auto_scale.setToolTip("Automatically adjust the scale by removing outliter data points.")
        
        # Add to the layout
        spacer1 = QSpacerItem(40, 20, QSizePolicy.Expanding, QSizePolicy.Minimum)
        combobox_layout.addWidget(self.map_type_label)
        combobox_layout.addWidget(self.cbb_map_type)
        combobox_layout.addWidget(self.cbb_palette)
        combobox_layout.addWidget(self.cb_auto_scale)
        combobox_layout.setContentsMargins(5, 5, 5, 5)

        # Add the combobox layout below the profile layout
        self.map_widget_layout.addLayout(combobox_layout)
        
        #### CREATE range sliders
        self.create_range_sliders(0,100)

        #### EXTRACT profil from 2Dmap
        profile_layout = QHBoxLayout()
        self.profile_name = QLineEdit(self)
        self.profile_name.setText("Profile_1")
        self.profile_name.setPlaceholderText("Profile_name...")
        self.profile_name.setFixedWidth(150)

        self.btn_extract_profile = QPushButton("Extract profil", self)
        self.btn_extract_profile.setToolTip("Extract profile data and plot it in Visu tab")
        self.btn_extract_profile.setFixedWidth(100)
        
        profile_layout.addWidget(self.profile_name)
        profile_layout.addWidget(self.btn_extract_profile)

        # Create Options Menu
        self.create_options_menu()
        
        self.tool_btn_options = QToolButton(self)
        self.tool_btn_options.setText("... ")
        self.tool_btn_options.setPopupMode(QToolButton.InstantPopup) 
        self.tool_btn_options.setMenu(self.options_menu) 
        spacer2 = QSpacerItem(40, 20, QSizePolicy.Expanding, QSizePolicy.Minimum)
        profile_layout.addItem(spacer2)
        profile_layout.addWidget(self.tool_btn_options)

        #### ADD PROFIL
        self.map_widget_layout.addLayout(profile_layout)
        profile_layout.setContentsMargins(5, 5, 5, 5)

    def update_settings(self):
        """Save selected wafer size to settings"""
        map_type = self.cbb_map_type.currentText()
        self.app_settings.map_type = map_type
        self.app_settings.save()

        #self.settings.setValue("map_type", map_type)
    
    def create_options_menu(self):
        """Create more view options for 2Dmap plot"""
        
        self.options_menu = QMenu(self)
        # Smoothing option
        options = [ ("Smoothing", "Smoothing", False),]
        for option_name, option_label, *checked in options:
            action = QAction(option_label, self)
            action.setCheckable(True)
            action.setChecked(checked[0] if checked else False)
            action.triggered.connect(self.refresh_plot)
            self.menu_actions[option_name] = action
            self.options_menu.addAction(action)  
        
    def create_range_sliders(self, xmin, xmax):
        """Create xrange and intensity-range sliders"""
        # Create x-axis range slider
        self.x_range_slider = QLabeledDoubleRangeSlider(Qt.Orientation.Horizontal)
        self.x_range_slider.setEdgeLabelMode(QLabeledDoubleRangeSlider.EdgeLabelMode.LabelIsValue)
        self.x_range_slider.setHandleLabelPosition(QLabeledDoubleRangeSlider.LabelPosition.NoLabel)
        self.x_range_slider.setSingleStep(0.01)
        self.x_range_slider.setRange(xmin, xmax)  
        self.x_range_slider.setValue((xmin, xmax)) 
        self.x_range_slider.setTracking(True)
        self.x_range_slider.valueChanged.connect(self.update_z_range_slider)

        self.x_range_slider_label = QLabel('X-range :')
        self.x_range_slider_label.setFixedWidth(80)  
        self.x_range_slider_label.setToolTip("Define the spectral range (X-range) for 'Area' and 'Maximum Intensity' calculation.")

        self.x_slider_layout = QHBoxLayout()
        self.x_slider_layout.addWidget(self.x_range_slider_label)
        self.x_slider_layout.addWidget(self.x_range_slider)
        
        self.x_slider_layout.setContentsMargins(5, 0, 5, 0)

        # Create z-axis range slider
        self.z_range_slider = QLabeledDoubleRangeSlider(Qt.Orientation.Horizontal)  
        self.z_range_slider.setEdgeLabelMode(QLabeledDoubleRangeSlider.EdgeLabelMode.LabelIsValue)
        self.z_range_slider.setHandleLabelPosition(QLabeledDoubleRangeSlider.LabelPosition.NoLabel)
        self.z_range_slider.setSingleStep(0.01)
        self.z_range_slider.setRange(0, 100) 
        self.z_range_slider.setValue((0, 100)) 
        self.z_range_slider.setTracking(True)    

        self.z_values_cbb = QComboBox()
        self.z_values_cbb.addItems(['Area', 'Max intensity']) 
        self.z_values_cbb.setFixedWidth(80)  
        self.z_values_cbb.setToolTip("Select parameter to plot 2Dmap")
        self.z_values_cbb.currentIndexChanged.connect(self.update_z_range_slider)
        self.z_range_slider.valueChanged.connect(self.refresh_plot)

        self.z_slider_layout = QHBoxLayout()
        self.z_slider_layout.addWidget(self.z_values_cbb)
        self.z_slider_layout.addWidget(self.z_range_slider)
        self.z_slider_layout.setContentsMargins(5, 0, 5, 0)
        
        self.map_widget_layout.addLayout(self.z_slider_layout)
        self.map_widget_layout.addLayout(self.x_slider_layout)

        vspacer = QSpacerItem(40, 20, QSizePolicy.Minimum, QSizePolicy.Expanding)
        self.map_widget_layout.addItem(vspacer)
    
    def populate_z_values_cbb(self):
        self.z_values_cbb.clear() 
        self.z_values_cbb.addItems(['Area', 'Max Intensity'])
        if not self.df_fit_results.empty:
            fit_columns = [col for col in self.df_fit_results.columns if col not in ['Filename', 'X', 'Y']]
            self.z_values_cbb.addItems(fit_columns)
  
    def refresh_plot(self):
        """Call the refresh_gui method of the main application."""
        if hasattr(self.main_app, 'refresh_gui'):
            
            self.main_app.refresh_gui()
        else:
            return
    
    def update_xrange_slider(self, xmin, xmax):
        """Update the range of the slider based on new min and max values."""
        xmin_label = round(xmin, 0)
        xmax_label = round(xmax, 0)
        self.x_range_slider.setRange(xmin, xmax)
        self.x_range_slider.setValue((xmin, xmax))
        #self.x_range_label.setText(f'[{xmin_label}; {xmax_label}]')
    
    def update_z_range_slider(self):
        if self.z_values_cbb.count() > 0 and self.z_values_cbb.currentIndex() >= 0:
            _,_, vmin, vmax, _ =self.get_data_for_heatmap()
            self.z_range_slider.setRange(vmin, vmax)
            self.z_range_slider.setValue((vmin, vmax))
        else:
            return
    
    def get_data_for_heatmap(self, map_type='2Dmap'):
        """Prepare data for heatmap based on range sliders values"""

        # Default return values in case of no valid map_df or filtered columns
        heatmap_pivot = pd.DataFrame()  # Empty DataFrame for heatmap
        extent = [0, 0, 0, 0]  # Default extent values
        vmin = 0
        vmax = 0
        grid_z = None # for waferplot
        
        if self.map_df is not None:
            xmin, xmax = self.x_range_slider.value()
            column_labels = self.map_df.columns[2:-1]  # Keep labels as strings

            # Convert slider range values to strings for comparison
            filtered_columns = column_labels[(column_labels.astype(float) >= xmin) &
                                            (column_labels.astype(float) <= xmax)]
            
            if len(filtered_columns) > 0:
                # Create a filtered DataFrame including X, Y, and the selected range of columns
                filtered_map_df = self.map_df[['X', 'Y'] + list(filtered_columns)]
                x_col = filtered_map_df['X'].values
                y_col = filtered_map_df['Y'].values
                final_z_col = []

                parameter = self.z_values_cbb.currentText()
                if parameter == 'Area':
                    # Intensity sums of of each spectrum over the selected range
                    z_col = filtered_map_df[filtered_columns].replace([np.inf, -np.inf], np.nan).fillna(0).clip(lower=0).sum(axis=1)
                elif parameter == 'Max Intensity':
                    # Max intensity value of each spectrum over the selected range
                    z_col = filtered_map_df[filtered_columns].replace([np.inf, -np.inf], np.nan).fillna(0).clip(lower=0).max(axis=1)
                else:
                    if not self.df_fit_results.empty:
                        map_name = self.map_df_name
                        # Plot only selected wafer/2Dmap
                        filtered_df = self.df_fit_results.query('Filename == @map_name')
                        if not filtered_df.empty and parameter in filtered_df.columns:
                            z_col = filtered_df[parameter]
                        else:
                            z_col = None
                
                # Auto scale 
                if self.cb_auto_scale.isChecked():
                    # Remove outliers using IQR method and replace them with interpolated values
                    Q1 = z_col.quantile(0.05)
                    Q3 = z_col.quantile(0.95)
                    IQR = Q3 - Q1
                    # Identify the outliers
                    outlier_mask = (z_col < (Q1 - 1.5 * IQR)) | (z_col > (Q3 + 1.5 * IQR))
                    # Interpolate values for the outliers using linear interpolation
                    z_col_interpolated = z_col.copy()
                    z_col_interpolated[outlier_mask] = np.nan  # Mark outliers as NaN for interpolation
                    z_col_interpolated = z_col_interpolated.interpolate(method='linear', limit_direction='both')
                    final_z_col=z_col_interpolated
                else:
                    final_z_col=z_col  

                try:
                    vmin = round(final_z_col.min(), 0)
                    vmax = round(final_z_col.max(), 0)
                except Exception as e:
                    #When the selected 'parameters' does not exist for selected wafer.
                    self.z_values_cbb.setCurrentIndex(0)

                    vmin = round(final_z_col.min(), 0)
                    vmax = round(final_z_col.max(), 0)

                self.number_of_points = len(set(zip(x_col, y_col)))
                
                if map_type != '2Dmap' and self.number_of_points >= 4:
                    # Create meshgrid for WaferPlot
                    r = self.get_wafer_radius(map_type)
                    grid_x, grid_y = np.meshgrid(np.linspace(-r, r, 300), np.linspace(-r, r, 300))
                    grid_z = griddata((x_col, y_col), final_z_col, (grid_x, grid_y), method='linear')
                    extent = [-r - 1, r + 1, -r - 0.5, r + 0.5]

                else:
                    # Regular 2D map
                    heatmap_data = pd.DataFrame({'X': x_col, 'Y': y_col, 'Z': z_col})
                    heatmap_pivot = heatmap_data.pivot(index='Y', columns='X', values='Z')
                    xmin, xmax = x_col.min(), x_col.max()
                    ymin, ymax = y_col.min(), y_col.max()
                    extent = [xmin, xmax, ymin, ymax]
                
        return heatmap_pivot, extent, vmin, vmax, grid_z
    
    def get_wafer_radius(self, map_type_text):
        """
        Extract wafer diameter from the map type string and return radius.
        Returns None if not a wafer type.
        """
        match = re.search(r'Wafer_(\d+)mm', map_type_text)
        if match:
            diameter = int(match.group(1))
            return diameter / 2
        return None
    
    def plot(self, coords):
        """Plot 2D maps of measurement points"""
        map_type = self.cbb_map_type.currentText()
        r = self.get_wafer_radius(map_type)
        self.ax.clear()
        
        # Plot wafer map
        if map_type != '2Dmap':
            wafer_circle = patches.Circle((0, 0), radius=r, fill=False,
                                        color='black', linewidth=1)
            self.ax.add_patch(wafer_circle)
            self.ax.set_yticklabels([])

            all_x, all_y = self.get_mes_sites_coord()
            self.ax.scatter(all_x, all_y, marker='x', color='gray', s=15)
            self.ax.grid(True, linestyle='--', linewidth=0.5, color='gray')
            
            
        heatmap_pivot, extent, vmin, vmax, grid_z  = self.get_data_for_heatmap(map_type)
        color = self.cbb_palette.currentText()
        interpolation_option = 'bilinear' if self.menu_actions['Smoothing'].isChecked() else 'none'
        vmin, vmax = self.z_range_slider.value()
    

        if map_type != '2Dmap' and self.number_of_points >= 4:
            self.img = self.ax.imshow(grid_z, extent=[-r - 0.5, r + 0.5, -r - 0.5, r + 0.5],
                            origin='lower', aspect='equal', cmap=color, interpolation='nearest')
            
        else: 
            self.img = self.ax.imshow(heatmap_pivot, extent=extent, vmin=vmin, vmax=vmax,
                            origin='lower', aspect='equal', cmap=color, interpolation=interpolation_option)
            #print(f'2Dmap: {heatmap_pivot}')
        
        # COLORBAR
        if hasattr(self, 'cbar') and self.cbar is not None:
            self.cbar.update_normal(self.img)
        else:
            self.cbar = self.ax.figure.colorbar(self.img, ax=self.ax)

        # MEASUREMENT SITES
        if coords:
            x, y = zip(*coords)
            self.ax.scatter(x, y, marker='o', color='red', s=20)

            #if map_type == '2Dmap':   
            self.plot_height_profile_on_map(coords)

        title = self.z_values_cbb.currentText()
        self.ax.set_title(title, fontsize=13)
        self.ax.get_figure().tight_layout()
        self.canvas.draw_idle()

    def plot_height_profile_on_map(self, coords):
        """Plot height profile directly on heatmap"""
        if len(coords) == 2:
            x, y = zip(*coords)
            self.ax.plot(x, y, color='black', linestyle='dotted', linewidth=2)

            # Extract profile values from the heatmap
            profile_df = self.extract_profile()
            if profile_df is not None:
                # Calculate the diagonal length of the heatmap
                extent = self.get_data_for_heatmap()[1]
                diagonal_length = np.sqrt((extent[1] - extent[0])**2 + (extent[3] - extent[2])**2)
                max_normalized_value = 0.3 * diagonal_length
                
                # Calculate height values
                profile_df['height'] = (profile_df['values'] / profile_df['values'].max()) * max_normalized_value
                profile_df['height'] -= profile_df['height'].min()

                x_vals = []
                y_vals = []

                # Calculate perpendicular points
                for i, row in profile_df.iterrows():
                    x_val = row['X']
                    y_val = row['Y']
                    normalized_distance = row['height']

                    # Calculate the direction vector of the line connecting the two points
                    dx = x[1] - x[0]
                    dy = y[1] - y[0]
                    length = np.sqrt(dx**2 + dy**2)

                    # Normalize the direction vector
                    if length > 0:
                        dx /= length
                        dy /= length

                    # Calculate the perpendicular direction
                    perp_dx = -dy
                    perp_dy = dx

                    # Calculate the new x and y values
                    x_vals.append(x_val + perp_dx * normalized_distance)
                    y_vals.append(y_val + perp_dy * normalized_distance)

                # Plot the height profile 
                self.ax.plot(x_vals, y_vals, color='black', linestyle='-', lw=2)

        
    def on_left_click_2Dmap(self, event):
        """select the measurement points via 2Dmap plot"""
        all_x, all_y = self.get_mes_sites_coord()
        self.spectra_listbox.clearSelection()
        if event.inaxes == self.ax:
            x_clicked, y_clicked = event.xdata, event.ydata
            if event.button == 1:  # Left mouse button
                all_x = np.array(all_x)
                all_y = np.array(all_y)
                distances = np.sqrt(
                    (all_x - x_clicked) ** 2 + (all_y - y_clicked) ** 2)
                nearest_index = np.argmin(distances)
                nearest_x, nearest_y = all_x[nearest_index], all_y[
                    nearest_index]

                # Check if Ctrl key is pressed
                modifiers = QApplication.keyboardModifiers()
                if modifiers == Qt.ControlModifier:
                    self.selected_points.append((nearest_x, nearest_y))
                else:
                    # Clear the selected points list and add the current one
                    self.selected_points = [(nearest_x, nearest_y)]

        # Set the current selection in the spectra_listbox
        for index in range(self.spectra_listbox.count()):
            item = self.spectra_listbox.item(index)
            item_text = item.text()
            x, y = map(float, item_text.strip('()').split(','))
            if (x, y) in self.selected_points:
                item.setSelected(True)
                self.current_row= index
                self.spectra_listbox.setCurrentRow(self.current_row)
            else:
                item.setSelected(False) 
            
    def extract_profile(self):
        """Extract a profile from 2D map plot via interpolation."""
        # Ensure exactly two points have been selected
        if len(self.selected_points) != 2:
            print("Select 2 points on map plot to define a profile")
            return None
        (x1, y1), (x2, y2) = self.selected_points
        heatmap_pivot, _, _, _, _ = self.get_data_for_heatmap()
        # Extract X and Y coordinates of the heatmap grid
        x_values = heatmap_pivot.columns.values
        y_values = heatmap_pivot.index.values
        z_values = heatmap_pivot.values
        # Interpolate Z values at the sampled points along the profile
        interpolator = RegularGridInterpolator((y_values, x_values), z_values)
        num_samples = 100 
        x_samples = np.linspace(x1, x2, num_samples)
        y_samples = np.linspace(y1, y2, num_samples)
        sample_points = np.vstack((y_samples, x_samples)).T 
        z_samples = interpolator(sample_points)
        # Calculate the distance from (x1, y1) to each sample point
        dists_from_start = np.sqrt((x_samples - x1)**2 + (y_samples - y1)**2)

        profile_df = pd.DataFrame({'X': x_samples, 'Y': y_samples, 'distance': dists_from_start,'values': z_samples})

        self.canvas.draw_idle()
        return profile_df
            
    def on_key_press(self, event):
        """Handler function for key press event"""
        if event.key == 'ctrl':
            self.ctrl_pressed = True

    def on_key_release(self, event):
        """Handler function for key release event"""
        if event.key == 'ctrl':
            self.ctrl_pressed = False

    def get_mes_sites_coord(self):
        """
        Get all coordinates of measurement sites of the selected map.
        """
        df = self.map_df
        all_x = df['X']
        all_y = df['Y']
        return all_x, all_y 

    def copy_fig(self):
        """Copy figure canvas to clipboard"""
        copy_fig_to_clb(self.canvas)

class CustomSpectra(Spectra):
    """Customized Spectra class of the fitspy package."""
    def apply_model(self, model_dict, fnames=None, ncpus=1,
                    show_progressbar=True):
        """ Apply 'model' to all or part of the spectra."""
        if fnames is None:
            fnames = self.fnames

        spectra = []
        for fname in fnames:
            spectrum, _ = self.get_objects(fname)
            
            # Customize the model_dict for this spectrum
            custom_model = deepcopy(model_dict)
            if hasattr(spectrum, "correction_value"):
                custom_model["correction_value"] = spectrum.correction_value
            if hasattr(spectrum, "is_corrected"):
                custom_model["is_corrected"] = spectrum.is_corrected

            spectrum.set_attributes(custom_model)
            spectrum.fname = fname  # reassign the correct fname
            spectra.append(spectrum)

        self.pbar_index = 0

        queue_incr = Queue()
        args = (queue_incr, len(fnames), ncpus, show_progressbar)
        thread = Thread(target=self.progressbar, args=args)
        thread.start()

        if ncpus == 1:
            for spectrum in spectra:
                spectrum.preprocess()
                spectrum.fit()
                queue_incr.put(1)
        else:
            fit_mp(spectra, ncpus, queue_incr)

        thread.join()


class SpectraViewWidget(QWidget):
    """Class to manage the spectra view widget."""
    def __init__(self, main_app):
        super().__init__()
        self.main_app = main_app # To connect to a method of main app (refresh gui)
        self.sel_spectrums =None
        self.peak_model = 'Lorentzian'
        self.dpi = 80
        self.figure = None
        self.ax = None
        self.canvas = None
        self.toolbar = None
        self.zoom_pan_active = False
        self.menu_actions = {}

        self.dragging_peak = None
        self.drag_event_connection = None
        self.release_event_connection = None

        self.initUI()
        QApplication.instance().focusChanged.connect(self.on_focus_changed)

    def initUI(self):
        """Initialize the UI components."""
        self.create_plot_widget()

    def create_plot_widget(self):
        """Create or update canvas and toolbar for plotting in the GUI."""
        plt.style.use(PLOT_POLICY)

        if not self.figure:
            self.create_figure_canvas_and_toolbar()
            self.create_tool_buttons()

            self.create_options_menu()
            self.create_normalization_widgets()
            self.create_copy_and_legend_buttons()
            self.create_control_layout()

        self.update_plot_styles()

    def create_figure_canvas_and_toolbar(self):
        self.figure = plt.figure(dpi=self.dpi)
        self.ax = self.figure.add_subplot(111)
        self.canvas = FigureCanvas(self.figure)
        self.canvas.mpl_connect('button_press_event', self.on_mouse_click)

        self.canvas.mpl_connect("scroll_event", self.on_scroll)

        self.toolbar = NavigationToolbar2QT(self.canvas, self)
        self.toolbar.zoom()  # Default zoom active

        for action in self.toolbar.actions():
            if action.text() in ['Home', 'Save', 'Pan', 'Back', 'Forward', 'Subplots', 'Zoom']:
                action.setVisible(False)
    
    def create_tool_buttons(self):
        self.btn_rescale = QPushButton("", self)
        self.btn_rescale.setToolTip("Rescale")
        self.btn_rescale.setIcon(QIcon(os.path.join(ICON_DIR, "rescale.png")))
        self.btn_rescale.setIconSize(QSize(24, 24))
        self.btn_rescale.clicked.connect(self.rescale)

        self.btn_zoom = QToolButton(self)
        self.btn_zoom.setCheckable(True)
        self.btn_zoom.setAutoExclusive(True)
        self.btn_zoom.setToolTip("Zoom")
        self.btn_zoom.setIcon(QIcon(os.path.join(ICON_DIR, "zoom.png")))
        self.btn_zoom.setIconSize(QSize(24, 24))
        self.btn_zoom.setChecked(True)
        self.btn_zoom.toggled.connect(self.toggle_zoom_pan)

        self.btn_baseline = QToolButton(self)
        self.btn_baseline.setCheckable(True)
        self.btn_baseline.setAutoExclusive(True)
        self.btn_baseline.setToolTip("Baseline")
        self.btn_baseline.setIcon(QIcon(os.path.join(ICON_DIR, "baseline.png")))
        self.btn_baseline.setIconSize(QSize(24, 24))

        self.btn_peak = QToolButton(self)
        self.btn_peak.setCheckable(True)
        self.btn_peak.setAutoExclusive(True)
        self.btn_peak.setToolTip("Peak")
        self.btn_peak.setIcon(QIcon(os.path.join(ICON_DIR, "peak.png")))
        self.btn_peak.setIconSize(QSize(24, 24))

    def create_normalization_widgets(self):
        self.btn_norm = QToolButton(self)
        self.btn_norm.setCheckable(True)
        self.btn_norm.setAutoExclusive(False)
        self.btn_norm.setToolTip("Normalization")
        self.btn_norm.setIcon(QIcon(os.path.join(ICON_DIR, "norm.png")))
        self.btn_norm.setIconSize(QSize(24, 24))
        self.btn_norm.clicked.connect(self.refresh_plot)
        self.btn_norm.clicked.connect(self.rescale)

        self.norm_x_min = QLineEdit(self)
        self.norm_x_min.setFixedWidth(40)
        self.norm_x_min.setPlaceholderText("Xmin")
        self.norm_x_min.setToolTip("Type Xmin for normalization")

        self.norm_x_max = QLineEdit(self)
        self.norm_x_max.setFixedWidth(40)
        self.norm_x_max.setPlaceholderText("Xmax")
        self.norm_x_max.setToolTip("Type Xmax for normalization")

    def create_copy_and_legend_buttons(self):
        self.btn_legend = QToolButton(self)
        self.btn_legend.setCheckable(True)
        self.btn_legend.setAutoExclusive(False)
        self.btn_legend.setToolTip("Show legend")
        self.btn_legend.setIcon(QIcon(os.path.join(ICON_DIR, "legend.png")))
        self.btn_legend.setIconSize(QSize(24, 24))
        self.btn_legend.clicked.connect(self.refresh_plot)

        self.btn_copy = QPushButton("", self)
        text = "Copy figure to clipboard.\nHold Ctrl & click to copy RAW & fitted curves to clipboard"
        self.btn_copy.setToolTip(text)
        self.btn_copy.setIcon(QIcon(os.path.join(ICON_DIR, "copy.png")))
        self.btn_copy.setIconSize(QSize(24, 24))
        self.btn_copy.clicked.connect(self.copy_fnc)

        self.R2 = QLabel("R2=0", self)

        self.tool_btn_options = QToolButton(self)
        self.tool_btn_options.setText("More options ")
        self.tool_btn_options.setPopupMode(QToolButton.InstantPopup)
        self.tool_btn_options.setMenu(self.options_menu)

    def create_control_layout(self):
        self.control_widget = QWidget(self)
        self.control_layout = QHBoxLayout(self.control_widget)
        self.control_layout.setContentsMargins(0, 0, 0, 0)

        self.control_layout.addWidget(self.btn_rescale)
        self.control_layout.addSpacing(10)
        self.control_layout.addWidget(self.btn_zoom)
        self.control_layout.addWidget(self.btn_baseline)
        self.control_layout.addWidget(self.btn_peak)
        self.control_layout.addSpacing(20)

        self.control_layout.addWidget(self.btn_norm)
        self.control_layout.addWidget(self.norm_x_min)
        self.control_layout.addWidget(self.norm_x_max)
        self.control_layout.addSpacing(20)

        self.control_layout.addWidget(self.btn_legend)
        self.control_layout.addSpacing(20)
        self.control_layout.addWidget(self.tool_btn_options)
        self.control_layout.addWidget(self.btn_copy)
        self.control_layout.addWidget(self.toolbar)
        self.control_layout.addWidget(self.R2)

        self.control_widget.setLayout(self.control_layout)


    def toggle_zoom_pan(self, checked):
        """Toggle zoom and pan functionality for spectra plot based on tool button selection."""
        if self.btn_zoom.isChecked():
            self.zoom_pan_active = True
            self.toolbar.zoom()  # Activate the zoom feature
        else:
            self.zoom_pan_active = False
            self.toolbar.zoom()  # Deactivate the zoom feature
            
    def on_scroll(self, event):
        ax = event.inaxes
        if ax is None:
            return  # Ignore scrolls outside the plot area

        y_min, y_max = ax.get_ylim()
        dy = (y_max - y_min) * 0.1  # 10% zoom step

        if event.step < 0:
            # Scroll up: increase max Y
            y_max = y_max + dy
        elif event.step > 0:
            # Scroll down: decrease max Y
            y_max = max(y_min + 1e-6, y_max - dy)  # prevent collapse

        ax.set_ylim(y_min, y_max)
        self.refresh_plot()   

    def create_options_menu(self):
        """Create widget containing all view options."""
        self.options_menu = QMenu(self)
        
        # X axis unit combobox
        xaxis_unit_widget = QWidget(self.options_menu)
        xaxis_unit_layout = QHBoxLayout(xaxis_unit_widget)
        xaxis_unit_label = QLabel("X-axis unit:", xaxis_unit_widget)
        xaxis_unit_layout.addWidget(xaxis_unit_label)

        self.cbb_xaxis_unit = QComboBox(xaxis_unit_widget)
        self.cbb_xaxis_unit.addItems(X_AXIS_UNIT)
        self.cbb_xaxis_unit.currentIndexChanged.connect(self.refresh_plot)
        xaxis_unit_layout.addWidget(self.cbb_xaxis_unit)
        xaxis_unit_layout.setContentsMargins(5, 5, 5, 5)
        
        # Create a QWidgetAction to hold the combined QLabel and QComboBox
        combo_action = QWidgetAction(self)
        combo_action.setDefaultWidget(xaxis_unit_widget)
        self.options_menu.addAction(combo_action)
        
        # Y axis scale
        yaxis_scale_widget = QWidget(self.options_menu)
        yaxis_scale_layout = QHBoxLayout(yaxis_scale_widget)
        yaxis_scale_label = QLabel("Y-axis scale:", yaxis_scale_widget)
        yaxis_scale_layout.addWidget(yaxis_scale_label)

        self.cbb_yaxis_scale = QComboBox(yaxis_scale_widget)
        self.cbb_yaxis_scale.addItems(['Linear scale', 'Log scale'])
        self.cbb_yaxis_scale.currentIndexChanged.connect(self.refresh_plot)
        yaxis_scale_layout.addWidget(self.cbb_yaxis_scale)
        yaxis_scale_layout.setContentsMargins(5, 5, 5, 5)
        
         # Create a QWidgetAction to hold the combined QLabel and QComboBox
        combo_action2 = QWidgetAction(self)
        combo_action2.setDefaultWidget(yaxis_scale_widget)
        self.options_menu.addAction(combo_action2)
        

        # Add a separator to distinguish the combobox from checkable actions
        self.options_menu.addSeparator()

        # Define view options with checkable actions
        options = [
            ("Colors", "Colors", True),
            ("Peaks", "Show Peaks"),
            ("Bestfit", "Best Fit", True),
            ("Raw", "Raw data"),
            ("Residual", "Residual"),
        ]

        # Add actions to the menu
        for option_name, option_label, *checked in options:
            action = QAction(option_label, self)
            action.setCheckable(True)
            action.setChecked(checked[0] if checked else False)
            action.triggered.connect(self.refresh_plot)
            self.menu_actions[option_name] = action
            self.options_menu.addAction(action)

        # Entry boxes for figure ratio
        ratio_widget = QWidget(self.options_menu)
        ratio_layout = QHBoxLayout(ratio_widget)

        fig_size_label = QLabel("Copied figure size:", ratio_widget)
        self.width_entry = QLineEdit(ratio_widget)
        self.width_entry.setFixedWidth(30)
        self.width_entry.setText("5.5")

        self.height_entry = QLineEdit(ratio_widget)
        self.height_entry.setFixedWidth(30)
        self.height_entry.setText("4")

        ratio_layout.addWidget(fig_size_label)
        ratio_layout.addWidget(self.width_entry)
        ratio_layout.addWidget(self.height_entry)
        ratio_layout.setContentsMargins(5, 5, 5, 5)

        # Create a QWidgetAction to hold the ratio input fields
        ratio_action = QWidgetAction(self)
        ratio_action.setDefaultWidget(ratio_widget)
        self.options_menu.addAction(ratio_action)

    def update_plot_styles(self):
        """Apply styles and settings to the plot."""
        xlable = self.cbb_xaxis_unit.currentText()
        self.ax.set_xlabel(xlable)
        self.ax.set_ylabel("Intensity (a.u)")
        self.ax.grid(True, linestyle='--', linewidth=0.5, color='gray')

    
    def rescale(self):
        """Rescale the spectra plot to fit within the axes."""
        self.ax.autoscale()
        self.canvas.draw()

    
    def set_peak_model(self, model):
        """Set the peak model to be used when clicking on the plot."""
        self.peak_model = model

    def refresh_gui(self):
        """Call the refresh_gui method of the main application."""
        if hasattr(self.main_app, 'refresh_gui'):
            self.main_app.refresh_gui()
        else:
            return


    def plot(self, sel_spectrums):
        """Plot spectra or fit results in the figure canvas."""
        if not sel_spectrums:
            self.clear_plot()
            return
        self.sel_spectrums = sel_spectrums

        self.prepare_plot()

        for spectrum in self.sel_spectrums:
            self.plot_spectrum(spectrum)

        self.finalize_plot()

    def prepare_plot(self):
        """Prepare the plot area before plotting spectra."""
        # Save current xlim and ylim to maintain zoom/pan state
        xlim, ylim = self.ax.get_xlim(), self.ax.get_ylim()
        self.ax.clear() 

        # Restore xlim and ylim if they were changed
        if not xlim == ylim == (0.0, 1.0):
            self.ax.set_xlim(xlim)
            self.ax.set_ylim(ylim)

    def plot_spectrum(self, spectrum):
        """Plot a single spectrum on the canvas."""
        x_values = spectrum.x
        y_values = self.get_y_values(spectrum)

        self.ax.plot(x_values, y_values, label=f"{spectrum.fname}", ms=3, lw=2)
        plot_baseline_dynamically(ax=self.ax, spectrum=spectrum)

        if self.menu_actions['Raw'].isChecked():
            self.plot_raw_data(spectrum)

        if self.menu_actions['Bestfit'].isChecked():
            self.plot_peaks_and_bestfit(spectrum)

        if self.menu_actions['Residual'].isChecked() and hasattr(spectrum.result_fit, 'residual'):
            try:
                self.plot_residual(spectrum)
            except:
                print("plot residual is not succesful")
            
        
        if hasattr(spectrum.result_fit, 'rsquared'):
            self.show_R2(spectrum)
        else:
            self.show_R2(None)

        # Reset color cycle if Colors option is not checked
        if not self.menu_actions['Colors'].isChecked():
            self.ax.set_prop_cycle(None)

    def get_y_values(self, spectrum):
        """Get y-values for a spectrum, applying normalization if needed."""
        x_values = spectrum.x
        y_values = spectrum.y

        if self.btn_norm.isChecked():
            norm_x_min = self.norm_x_min.text().strip()
            norm_x_max = self.norm_x_max.text().strip() 

            if norm_x_min and norm_x_max:  # If user provided both X min and X max values
                try:
                    norm_x_min = float(norm_x_min)
                    norm_x_max = float(norm_x_max)
                    # Ensure min is less than max
                    if norm_x_min > norm_x_max:
                        norm_x_min, norm_x_max = norm_x_max, norm_x_min 
                    
                    # Find the closest indices in x_values
                    min_index = (np.abs(x_values - norm_x_min)).argmin()
                    max_index = (np.abs(x_values - norm_x_max)).argmin()

                    # Get max Y value within the range
                    norm_y_value = max(y_values[min_index:max_index + 1])
                except ValueError:
                    print("Invalid X value. Normalizing to max intensity instead.")
                    norm_y_value = max(y_values) 
            else:
                norm_y_value = max(y_values)  

            if norm_y_value != 0:
                y_values = y_values / norm_y_value 
        return y_values


    def plot_raw_data(self, spectrum):
        """Plot raw data points if the option is checked."""
        x0_values = spectrum.x0
        y0_values = spectrum.y0
        self.ax.plot(x0_values, y0_values, 'ko-', label='raw', ms=3, lw=1)

    def plot_peak(self, y_peak, x_values, peak_label, peak_model):
        """Plot individual peak, optionally filled, and return line and peak info."""
        
        line, = self.ax.plot(x_values, y_peak, '-', label=peak_label, lw=1.5)
        
        # Annotate if enabled
        if self.menu_actions['Peaks'].isChecked():
            self.annotate_peak(peak_model, peak_label)

        # Extract peak info for hover and interaction
        peak_info = {
            "peak_label": peak_label,
            "peak_model": peak_model, # For peak dragging features.
        }

        # Extract parameter values (x0, fwhm, amplitude, etc.)
        if hasattr(peak_model, 'param_names') and hasattr(peak_model, 'param_hints'):
            for param_name in peak_model.param_names:
                key = param_name.split('_', 1)[1]  # e.g., x0, amplitude
                if key in peak_model.param_hints and 'value' in peak_model.param_hints[key]:
                    val = peak_model.param_hints[key]['value']
                    peak_info[key] = val
        return line, peak_info

    def plot_peaks_and_bestfit(self, spectrum):
        x_values = spectrum.x
        y_peaks = np.zeros_like(x_values)
        y_bkg = self.get_background_y_values(spectrum)

        peak_labels = spectrum.peak_labels
        self.fitted_lines = []  # Store line and info for hover

        for i, peak_model in enumerate(spectrum.peak_models):
            y_peak = self.evaluate_peak_model(peak_model, x_values)
            y_peaks += y_peak
            result = self.plot_peak(y_peak, x_values, peak_labels[i], peak_model)

            if result is not None:
                line, peak_info = result
                if line is not None:
                    self.fitted_lines.append((line, peak_info))
                else:
                    pass
            else:
                pass

        if hasattr(spectrum.result_fit, 'success'):
            y_fit = y_bkg + y_peaks
            self.ax.plot(x_values, y_fit, label="bestfit")

        self.enable_hover_highlight()  # connect hover after drawing lines
    
    def enable_hover_highlight(self):
        if not hasattr(self, 'hover_connection'):
            self.hover_connection = self.canvas.mpl_connect('motion_notify_event', self.on_hover)

    def on_hover(self, event):
        if event.inaxes != self.ax or not self.canvas.isActiveWindow():
            self.hide_tooltip()
            return

        for line, info in self.fitted_lines:
            if line.contains(event)[0]:
                # Define the keys we want to show, in order
                fields = [
                    ('label', info.get('peak_label')),
                    ('center', info.get('x0')),
                    ('intensity', info.get('ampli')),
                    ('fwhm', info.get('fwhm')),
                    ('fwhm_l', info.get('fwhm_l')),
                    ('fwhm_r', info.get('fwhm_r')),
                    ('alpha', info.get('alpha')),
                ]

                # Build the tooltip string dynamically
                lines = []
                for label, val in fields:
                    if val is not None:
                        try:
                            val_str = f"{val:.3f}" if isinstance(val, (float, int)) else str(val)
                        except Exception:
                            val_str = str(val)
                        lines.append(f"{label}: {val_str}")

                text = "\n".join(lines)
                self.show_tooltip(event, text)
                self._highlight_line(line)

                # Connect mouse press for dragging
                self.canvas.mpl_disconnect(getattr(self, 'click_connection', None))
                self.click_connection = self.canvas.mpl_connect('button_press_event', self.on_mouse_click)
                return

        self.hide_tooltip()
        self._reset_highlight()
    
    def on_mouse_click(self, event):
        """interaction with peak model and background via left-right mouse click"""
        if event.inaxes != self.ax or not self.sel_spectrums:
            return

        if self.zoom_pan_active:
            return

        x_click = event.xdata
        y_click = event.ydata
        sel_spectrum = self.sel_spectrums[0]

        if self.btn_peak.isChecked():
            if event.button == 1:
                # Try to drag peak if hovered over a line
                for line, info in self.fitted_lines:
                    if line.contains(event)[0]:
                        self.dragging_peak = (line, info)
                        self.drag_event_connection = self.canvas.mpl_connect('motion_notify_event', self.on_drag_peak)
                        self.release_event_connection = self.canvas.mpl_connect('button_release_event', self.on_release_drag)
                        return  # do not add a new peak if we start dragging

                # Else, normal left-click to add peak
                sel_spectrum.add_peak_model(self.peak_model, x_click, dfwhm=200)
                self.refresh_gui()

            elif event.button == 3:
                # Right-click: remove closest peak
                if hasattr(sel_spectrum, "peak_models") and sel_spectrum.peak_models:
                    closest_idx = min(
                        range(len(sel_spectrum.peak_models)),
                        key=lambda i: abs(sel_spectrum.peak_models[i].param_hints['x0']['value'] - x_click)
                    )
                    del sel_spectrum.peak_models[closest_idx]
                    del sel_spectrum.peak_labels[closest_idx]
                    self.refresh_gui()

        elif self.btn_baseline.isChecked():
            if event.button == 1:
                if sel_spectrum.baseline.is_subtracted:
                    show_alert("Baseline is already subtracted. Reinitialize spectrum to perform new baseline")
                else:
                    sel_spectrum.baseline.add_point(x_click, y_click)
                self.refresh_gui()

            elif event.button == 3:
                if (hasattr(sel_spectrum.baseline, "points") and
                    isinstance(sel_spectrum.baseline.points, list) and
                    len(sel_spectrum.baseline.points[0]) > 0):
                    x_points = sel_spectrum.baseline.points[0]
                    y_points = sel_spectrum.baseline.points[1]
                    closest_idx = min(range(len(x_points)), key=lambda i: abs(x_points[i] - x_click))
                    x_points.pop(closest_idx)
                    y_points.pop(closest_idx)
                    self.refresh_gui()                
    
    def show_tooltip(self, event, text):
        if not hasattr(self, 'tooltip'):
            from PySide6.QtWidgets import QLabel
            self.tooltip = QLabel(self.canvas)

            self.tooltip.setStyleSheet("""
                background-color: rgba(255, 255, 255, 0.5);
                color: black;
                border: 0.1px gray;
                padding: 2px;
            """)
            self.tooltip.setWindowFlags(Qt.ToolTip)

        self.tooltip.setText(text)

        cursor_pos = QCursor.pos()
        offset = QPoint(5, -75)
        self.tooltip.move(cursor_pos + offset)
        self.tooltip.show()

    def on_focus_changed(self, old, new):
        if not self.canvas.isActiveWindow():
            self.hide_tooltip()
        
    def hide_tooltip(self):
        if hasattr(self, 'tooltip'):
            self.tooltip.hide()

    def _highlight_line(self, line_to_highlight):
        """Highlight the peak upon hover mouse cursor"""
        # If already highlighted this line, do nothing
        if getattr(self, 'highlighted_line', None) == line_to_highlight:
            return
        self._reset_highlight()

        # Save current linewidth to restore later
        line_to_highlight._orig_lw = line_to_highlight.get_linewidth()

        # Increase linewidth
        line_to_highlight.set_linewidth(3)
        self.highlighted_line = line_to_highlight

        # Redraw canvas to reflect changes
        self.canvas.draw_idle()

    def _reset_highlight(self):
        """Un-Highlight the peak upon hover mouse cursor"""
        if hasattr(self, 'highlighted_line') and self.highlighted_line is not None:
            # Restore original linewidth
            orig_lw = getattr(self.highlighted_line, '_orig_lw', 1.5)
            self.highlighted_line.set_linewidth(orig_lw)

            self.highlighted_line = None
            self.canvas.draw_idle()

    def on_drag_peak(self, event):
        """Dragging peak to adjust x0 of peak_model in real-time"""
        if self.dragging_peak is None or event.xdata is None:
            return

        line, info = self.dragging_peak
        peak_model = info.get('peak_model')
        if not peak_model:
            return

        # Update x0 in the model
        peak_model.param_hints['x0']['value'] = event.xdata

        # Re-plot this spectrum
        self.plot(self.sel_spectrums)
        

    def on_release_drag(self, event):
        self.dragging_peak = None
        if hasattr(self, 'drag_event_connection'):
            self.canvas.mpl_disconnect(self.drag_event_connection)
            self.drag_event_connection = None
        if hasattr(self, 'release_event_connection'):
            self.canvas.mpl_disconnect(self.release_event_connection)
            self.release_event_connection = None
        self.refresh_gui() 


    def get_background_y_values(self, spectrum):
        """Get y-values for the background model."""
        x_values = spectrum.x
        if spectrum.bkg_model is not None:
            return spectrum.bkg_model.eval(spectrum.bkg_model.make_params(), x=x_values)
        return np.zeros_like(x_values)

    def evaluate_peak_model(self, peak_model, x_values):
        """Evaluate the peak model to get y-values."""
        param_hints_orig = deepcopy(peak_model.param_hints)
        for key in peak_model.param_hints.keys():
            peak_model.param_hints[key]['expr'] = ''
        
        params = peak_model.make_params()
        peak_model.param_hints = param_hints_orig
        
        return peak_model.eval(params, x=x_values) 

    def annotate_peak(self, peak_model, peak_label):
        """Annotate peaks on the plot with labels."""
        position = peak_model.param_hints['x0']['value']
        intensity = peak_model.param_hints['ampli']['value']
        position = round(position, 2)
        text = f"{peak_label}\n({position})"
        self.ax.text(position, intensity, text, ha='center', va='bottom', color='black', fontsize=12)

    def compute_residual(self, spectrum):
        """Compute residual = raw data - (background + sum of peaks)."""
        x_values = spectrum.x
        y_values = spectrum.y
        y_bkg = self.get_background_y_values(spectrum)

        # Sum all peak models
        y_peaks = np.zeros_like(x_values)
        for peak_model in spectrum.peak_models:
            y_peak = self.evaluate_peak_model(peak_model, x_values)
            y_peaks += y_peak

        y_fit = y_bkg + y_peaks
        residual = y_values - y_fit
        return x_values, residual

    def plot_residual(self, spectrum):
        """Plot the residuals if available."""
        x_values, residual = self.compute_residual(spectrum)
        # x_values = spectrum.x
        # residual = spectrum.result_fit.residual  # Bug of fitspy 2025.6 version
        self.ax.plot(x_values, residual, 'ko-', ms=3, label='residual')

    def show_R2(self, spectrum):
        """Display R² value in the GUI."""
        if spectrum is not None and hasattr(spectrum.result_fit, 'rsquared'):
            rsquared = round(spectrum.result_fit.rsquared, 4)
            self.R2.setText(f"R²={rsquared}")
        else:
            self.R2.setText("R²=0")

    def finalize_plot(self):
        """Finalize plot settings and draw the canvas."""
        # Use the selected x-axis label from the combobox
        xlabel = self.cbb_xaxis_unit.currentText() if self.cbb_xaxis_unit else "Wavenumber (cm-1)"
        self.ax.set_xlabel(xlabel)
        self.ax.set_ylabel("Intensity (a.u)")
        y_scale = self.cbb_yaxis_scale.currentText()
        if y_scale == 'Log scale':
            self.ax.set_yscale('log')
        else:  # Default to linear scale
            self.ax.set_yscale('linear')

        if self.btn_legend.isChecked():
            self.ax.legend(loc='upper right')

        self.ax.grid(True, linestyle='--', linewidth=0.5, color='gray')
        self.figure.tight_layout()
        self.canvas.draw_idle()
        
    def clear_plot(self):
        """Explicitly clear the spectra plot."""
        if self.ax:
            self.ax.clear()
            self.ax.set_xlabel("X-axis")
            self.ax.set_ylabel("Y-axis")
            self.ax.grid(True, linestyle='--', linewidth=0.5, color='gray')
            self.canvas.draw_idle()  
            
    def refresh_plot(self):
        """Refresh the plot based on user view options."""
        if not self.sel_spectrums:
            self.clear_plot() 
        else:
            self.plot(self.sel_spectrums)

    def copy_fig(self):
        """Copy figure canvas to clipboard"""
        width_text = self.width_entry.text().strip()
        height_text = self.height_entry.text().strip()

        # Set default values if the entry boxes are empty
        width = float(width_text) if width_text else 5.5  # Default width
        height = float(height_text) if height_text else 4.0
        copy_fig_to_clb(self.canvas, size_ratio=(width, height))
        
    def copy_spectra_data(self):
        """Copy X, Y, and peak model data of the first selected spectrum to clipboard as a DataFrame."""
        import pandas as pd

        if not self.sel_spectrums or len(self.sel_spectrums) == 0:
            print("No spectrum selected.")
            return

        spectrum = self.sel_spectrums[0]
        x_values = spectrum.x
        y_values = spectrum.y

        # Create a dictionary for the DataFrame
        data = {
            "X values": x_values,
            "Y values": y_values
        }

        # Add each peak model’s evaluated Y values as a new column
        for i, peak_model in enumerate(spectrum.peak_models):
            y_peak = self.evaluate_peak_model(peak_model, x_values)

            if hasattr(spectrum, 'peak_labels') and i < len(spectrum.peak_labels):
                label = spectrum.peak_labels[i]
            else:
                label = f"Peak {i + 1}"

            data[label] = y_peak

        df = pd.DataFrame(data)
        df.to_clipboard(index=False)
        print("Spectrum data copied to clipboard.")
        
    def copy_fnc(self):
        modifiers = QApplication.keyboardModifiers()
        if modifiers == Qt.ControlModifier:
            self.copy_spectra_data()
        else:
            self.copy_fig()
        
        
class FilterWidget:
    """
    Class for Handling "Filter Features" in Querying Pandas DataFrames

    Attributes:
    line_edit (QLineEdit): Input field for filter expressions.
    listbox (QListWidget): List widget to display filter expressions as checkboxes.
    df (pandas.DataFrame): DataFrame to be filtered.
    filters (list): List to store filter expressions and their states.
    """

    def __init__(self, df):
        """
        Initialize the Filter class with a DataFrame and set up UI components.

        Args:
        df (pandas.DataFrame): The DataFrame to apply filters on.
        """
        self.df = df
        self.filters = []
        self.initUI()

    def initUI(self):
        """Initialize the UI components."""
        self.create_filter_widget()

    def create_filter_widget(self):
        """Create filter UI components and organize them directly within the QGroupBox."""
        # Create Group Box to hold all filter widgets
        self.gb_filter_widget = QGroupBox()
        self.gb_filter_widget.setTitle(QCoreApplication.translate("mainWindow", u"Data filtering:", None))

        # Set the main layout for the group box
        self.layout_main = QVBoxLayout(self.gb_filter_widget)

        # Horizontal layout to hold the filter entry and buttons
        self.layout_buttons = QHBoxLayout()
        self.layout_buttons.setSpacing(2)

        # Entry box for filter queries
        self.filter_query = QLineEdit(self.gb_filter_widget)
        self.filter_query.setPlaceholderText("Enter your filter expression...") 
        self.filter_query.returnPressed.connect(self.add_filter)
        self.layout_buttons.addWidget(self.filter_query)

        # Button to add a filter
        self.btn_add_filter = QPushButton(self.gb_filter_widget)
        icon_add = QIcon()
        icon_add.addFile(u":/icon/iconpack/add.png", QSize(), QIcon.Normal, QIcon.Off)
        self.btn_add_filter.setIcon(icon_add)
        self.btn_add_filter.clicked.connect(self.add_filter) 
        self.layout_buttons.addWidget(self.btn_add_filter)

        # Button to remove selected filters
        self.btn_remove = QPushButton(self.gb_filter_widget)
        icon_remove = QIcon()
        icon_remove.addFile(u":/icon/iconpack/close.png", QSize(), QIcon.Normal, QIcon.Off)
        self.btn_remove.setIcon(icon_remove)
        self.btn_remove.clicked.connect(self.remove_filter) 
        self.layout_buttons.addWidget(self.btn_remove)

        # Button to apply filters
        self.btn_apply = QPushButton(self.gb_filter_widget)
        icon_apply = QIcon()
        icon_apply.addFile(u":/icon/iconpack/done.png", QSize(), QIcon.Normal, QIcon.Off)
        self.btn_apply.setIcon(icon_apply)
        self.btn_apply.setText("Apply")  
        self.btn_apply.setToolTip("Click to apply checked filters to the selected dataframe") 
        self.btn_apply.clicked.connect(self.apply_filters)  
        self.layout_buttons.addWidget(self.btn_apply)

        # Add the horizontal layout to the main layout of the group box
        self.layout_main.addLayout(self.layout_buttons)

        # Create QListWidget to display filter expressions as checkboxes
        self.filter_listbox = QListWidget(self.gb_filter_widget)
        self.layout_main.addWidget(self.filter_listbox)

    def set_dataframe(self, df):
        """Set the DataFrame to be filtered."""
        self.df = df

    def add_filter(self):
        """Add a filter expression to the filters list and update the UI."""
        filter_expression = self.filter_query.text().strip()
        if filter_expression:
            filter = {"expression": filter_expression, "state": False}
            self.filters.append(filter)
        # Add the filter expression to QListWidget as a checkbox item
        item = QListWidgetItem()
        checkbox = QCheckBox(filter_expression)
        item.setSizeHint(checkbox.sizeHint())
        self.filter_listbox.addItem(item)
        self.filter_listbox.setItemWidget(item, checkbox)

    def remove_filter(self):
        """Remove selected filter(s) from the filters list and UI."""
        selected_items = [item for item in self.filter_listbox.selectedItems()]
        for item in selected_items:
            checkbox = self.filter_listbox.itemWidget(item)
            filter_expression = checkbox.text()
            for filter in self.filters[:]:
                if filter.get("expression") == filter_expression:
                    self.filters.remove(filter)
            self.filter_listbox.takeItem(self.filter_listbox.row(item))

    def get_current_filters(self):
        """
        Retrieve the current state of filters as displayed in the UI.

        Returns:
        list: List of dictionaries representing filter expressions and their states.
        Each dictionary has keys 'expression' and 'state'.
        """
        checked_filters = []
        for i in range(self.filter_listbox.count()):
            item = self.filter_listbox.item(i)
            checkbox = self.filter_listbox.itemWidget(item)
            expression = checkbox.text()
            state = checkbox.isChecked()
            checked_filters.append({"expression": expression, "state": state})
        return checked_filters

    def apply_filters(self, filters=None):
        """
        Apply filters to the DataFrame (self.df) based on the current or provided filters.

        Args:
        filters (list, optional): List of dictionaries representing filter expressions and their states.
                                  Defaults to None, meaning current UI filters are used.

        Returns:
        pandas.DataFrame or None: Filtered DataFrame based on applied filters or None if self.df is None.
        """
        if filters:
            self.filters = filters
        else:
            checked_filters = self.get_current_filters()
            self.filters = checked_filters

        # Apply all filters at once
        self.filtered_df = self.df.copy() if self.df is not None else None

        if self.filtered_df is not None:  # Check if filtered_df is not None
            for filter_data in self.filters:
                filter_expr = filter_data["expression"]
                is_checked = filter_data["state"]
                if is_checked:
                    try:
                        filter_expr = str(filter_expr)
                        print(f"Applying filter expression: {filter_expr}")
                        # Apply the filter
                        self.filtered_df = self.filtered_df.query(filter_expr)
                    except Exception as e:
                        print(f"Error applying filter: {str(e)}")

        return self.filtered_df

    def upd_filter_listbox(self):
        """
        Update the listbox UI to reflect changes in filters.

        Clears the listbox and re-populates it with current filters.
        Each filter is displayed as a checkbox item.
        """
        self.filter_listbox.clear()
        for filter_data in self.filters:
            filter_expression = filter_data["expression"]
            item = QListWidgetItem()
            checkbox = QCheckBox(filter_expression)
            item.setSizeHint(checkbox.sizeHint())
            self.filter_listbox.addItem(item)
            self.filter_listbox.setItemWidget(item, checkbox)
            checkbox.setChecked(filter_data["state"])
class PeakTableWidget:
    """Class dedicated to show fit parameters of Spectrum objects in the GUI"""

    def __init__(self, main_app, main_layout, cbb_layout):
        # the main app where the PeakTable class is implemented, so we can connect to the method of main-map (upd_spectra_list)
        self.main_app = main_app 
        self.main_layout = main_layout # layout where the peak_table are placed
        self.cbb_layout = cbb_layout  # layout where comboboxes are placed
        self.sel_spectrum = None

        # Initialize Checkboxes
        self.cb_limits = QCheckBox("Limits")
        self.cb_expr = QCheckBox("Expression")
        self.cbb_layout.addWidget(self.cb_limits)
        self.cbb_layout.addWidget(self.cb_expr)
        self.cb_limits.stateChanged.connect(self.refresh_gui)
        self.cb_expr.stateChanged.connect(self.refresh_gui)

    def clear_layout(self, layout):
        """To clear a given layout"""
        if layout is not None:
            while layout.count():
                item = layout.takeAt(0)
                widget = item.widget()
                if widget:
                    widget.deleteLater()
                else:
                    self.clear_layout(item.layout())

    def show(self, sel_spectrum=None):
        """To show all fitted parameters in GUI"""
        if sel_spectrum is None:
            self.clear()
            return
        
        self.sel_spectrum = sel_spectrum
        self.clear_layout(self.main_layout)
        header_labels = ["  ", "Label", "Model"]
        param_hint_order = ['x0', 'fwhm', 'fwhm_l', 'fwhm_r', 'ampli', 'alpha']

        # Create and add headers to list
        for param_hint_key in param_hint_order:
            if any(param_hint_key in peak_model.param_hints for peak_model in
                   self.sel_spectrum.peak_models):
                header_labels.append(param_hint_key.title())
                header_labels.append(f"fix {param_hint_key.title()}")
                if self.cb_limits.isChecked():
                    header_labels.append(f"min {param_hint_key.title()}")
                    header_labels.append(f"max {param_hint_key.title()}")
                if self.cb_expr.isChecked():
                    header_labels.append(f"expression {param_hint_key.title()}")

        # Create vertical layouts for each column type
        delete_layout = QVBoxLayout()
        label_layout = QVBoxLayout()
        model_layout = QVBoxLayout()
        param_hint_layouts = {param_hint: {var: QVBoxLayout() for var in ['value', 'min', 'max', 'expr','vary']} for
                              param_hint in param_hint_order}

        # Add header labels to each layout
        for header_label in header_labels:
            label = QLabel(header_label)
            label.setAlignment(Qt.AlignCenter)
            if header_label == "  ":
                delete_layout.addWidget(label)
            elif header_label == "Label":
                label_layout.addWidget(label)
            elif header_label == "Model":
                model_layout.addWidget(label)
            elif header_label.startswith("fix"):
                param_hint_key = header_label.split()[1].lower()
                param_hint_layouts[param_hint_key]['vary'].addWidget(label)
            elif "min" in header_label:
                param_hint_key = header_label.split()[1].lower()
                param_hint_layouts[param_hint_key]['min'].addWidget(label)
            elif "max" in header_label:
                param_hint_key = header_label.split()[1].lower()
                param_hint_layouts[param_hint_key]['max'].addWidget(label)
            elif "expression" in header_label:
                param_hint_key = header_label.split()[1].lower()
                param_hint_layouts[param_hint_key]['expr'].addWidget(label)
            else:
                param_hint_key = header_label.lower()
                param_hint_layouts[param_hint_key]['value'].addWidget(label)

        for i, peak_model in enumerate(self.sel_spectrum.peak_models):
            # Button to delete peak_model
            delete = QPushButton(peak_model.prefix)
            icon = QIcon()
            icon.addFile(os.path.join(ICON_DIR, "close.png"))
            delete.setIcon(icon)
            delete.setFixedWidth(50)
            delete.clicked.connect(self.delete_helper(self.sel_spectrum, i))
            delete_layout.addWidget(delete)

            # Peak_label
            label = QLineEdit(self.sel_spectrum.peak_labels[i])
            label.setFixedWidth(80)
            label.textChanged.connect(
                lambda text, idx=i, spectrum=self.sel_spectrum: self.update_peak_label(spectrum,idx, text))
            label_layout.addWidget(label)

            # Peak model : Lorentizan, Gaussian, etc...
            model = QComboBox()
            model.addItems(PEAK_MODELS)
            current_model_index = PEAK_MODELS.index(
                peak_model.name2) if peak_model.name2 in PEAK_MODELS else 0
            model.setCurrentIndex(current_model_index)
            model.setFixedWidth(120)
            model.currentIndexChanged.connect(
                lambda index, spectrum=self.sel_spectrum, idx=i,
                       combo=model: self.update_model_name(spectrum, index, idx, combo.currentText()))
            model_layout.addWidget(model)

            # variables of peak_model
            param_hints = peak_model.param_hints
            for param_hint_key in param_hint_order:
                if param_hint_key in param_hints:
                    param_hint_value = param_hints[param_hint_key]

                    # 4.1 VALUE
                    value_val = round(param_hint_value.get('value', 0.0), 3)
                    value = QLineEdit(str(value_val))
                    value.setFixedWidth(70)
                    value.setFixedHeight(24)
                    value.setAlignment(Qt.AlignRight)
                    value.textChanged.connect(lambda text, pm=peak_model, key=param_hint_key: self.update_param_hint_value(pm, key, text))
                    param_hint_layouts[param_hint_key]['value'].addWidget(value)

                    # 4.2 FIXED or NOT
                    vary = QCheckBox()
                    vary.setChecked(not param_hint_value.get('vary', False))
                    vary.setFixedHeight(24)

                    # Create container widget with horizontal layout to center the checkbox
                    checkbox_container = QWidget()
                    checkbox_layout = QHBoxLayout()
                    checkbox_layout.setContentsMargins(0, 0, 0, 0)
                    checkbox_layout.setAlignment(Qt.AlignCenter)
                    checkbox_layout.addWidget(vary)
                    checkbox_container.setLayout(checkbox_layout)

                    vary.stateChanged.connect(
                        lambda state, pm=peak_model,
                            key=param_hint_key: self.update_param_hint_vary(pm, key, not state))

                    param_hint_layouts[param_hint_key]['vary'].addWidget(checkbox_container)

                    # 4.3 MIN MAX
                    if self.cb_limits.isChecked():
                        min_val = round(param_hint_value.get('min', 0.0), 2)
                        min_lineedit = QLineEdit(str(min_val))
                        min_lineedit.setFixedWidth(70)
                        min_lineedit.setFixedHeight(24)
                        min_lineedit.setAlignment(Qt.AlignRight)
                        min_lineedit.textChanged.connect(
                            lambda text, pm=peak_model,key=param_hint_key:
                            self.update_param_hint_min(pm, key, text))
                        param_hint_layouts[param_hint_key]['min'].addWidget(min_lineedit)

                        max_val = round(param_hint_value.get('max', 0.0), 2)
                        max_lineedit = QLineEdit(str(max_val))
                        max_lineedit.setFixedWidth(70)
                        max_lineedit.setFixedHeight(24)
                        max_lineedit.setAlignment(Qt.AlignRight)
                        max_lineedit.textChanged.connect(
                            lambda text, pm=peak_model, key=param_hint_key:
                            self.update_param_hint_max(pm, key, text))
                        param_hint_layouts[param_hint_key]['max'].addWidget(max_lineedit)

                    # 4.4 EXPRESSION
                    if self.cb_expr.isChecked():
                        expr_val = str(param_hint_value.get('expr', ''))
                        expr = QLineEdit(expr_val)
                        expr.setFixedWidth(150)
                        expr.setFixedHeight(
                            24)  # Set a fixed height for QLineEdit
                        expr.setAlignment(Qt.AlignRight)
                        expr.textChanged.connect(
                            lambda text, pm=peak_model,
                                   key=param_hint_key:
                            self.update_param_hint_expr(
                                pm, key, text))
                        param_hint_layouts[param_hint_key]['expr'].addWidget(expr)
                else:
                    # Add empty labels for alignment
                    empty_label = QLabel()
                    empty_label.setFixedHeight(24)
                    param_hint_layouts[param_hint_key]['value'].addWidget(empty_label)
                    param_hint_layouts[param_hint_key]['vary'].addWidget(empty_label)
                    if self.cb_limits.isChecked():
                        param_hint_layouts[param_hint_key]['min'].addWidget(empty_label)
                        param_hint_layouts[param_hint_key]['max'].addWidget(empty_label)
                    if self.cb_expr.isChecked():
                        param_hint_layouts[param_hint_key]['expr'].addWidget(empty_label)

        # Add vertical layouts to main layout
        self.main_layout.addLayout(delete_layout)
        self.main_layout.addLayout(label_layout)
        self.main_layout.addLayout(model_layout)

        for param_hint_key, param_hint_layout in param_hint_layouts.items():
            for var_layout in param_hint_layout.values():
                self.main_layout.addLayout(var_layout)
                
        # Add a horizontal spacer to absorb any remaining space
        spacer = QSpacerItem(20, 20, QSizePolicy.Expanding, QSizePolicy.Minimum)
        self.main_layout.addItem(spacer)

    def update_model_name(self, spectrum, index, idx, new_model):
        """ Update the model function (Lorentizan, Gaussian...) related to
        the ith-model """
        old_model_name = spectrum.peak_models[idx].name2
        new_model_name = new_model
        if new_model_name != old_model_name:
            ampli = spectrum.peak_models[idx].param_hints['ampli']['value']
            x0 = spectrum.peak_models[idx].param_hints['x0']['value']
            peak_model = spectrum.create_peak_model(idx + 1, new_model_name, x0=x0, ampli=ampli, dx0=(20., 20.))
            spectrum.peak_models[idx] = peak_model
            spectrum.result_fit = lambda: None
            self.refresh_gui()  # To update in GUI of main application.

    def delete_helper(self, spectrum, idx):
        """Helper method"""
        return lambda: self.delete_peak_model(spectrum, idx)
    
    def delete_peak_model(self, spectrum, idx):
        """To delete a peak model"""
        del spectrum.peak_models[idx]
        del spectrum.peak_labels[idx]
        self.refresh_gui()  # To update in GUI of main application.
        
    def refresh_gui(self):
        """Call the refresh_gui method of the main application."""
        if hasattr(self.main_app, 'refresh_gui'):
            self.main_app.refresh_gui()
        else:
            print("Main application does not have upd_spectra_list method.")

    def update_peak_label(self, spectrum, idx, text):
        spectrum.peak_labels[idx] = text

    def update_param_hint_value(self, pm, key, text):
        pm.param_hints[key]['value'] = float(text)

    def update_param_hint_min(self, pm, key, text):
        pm.param_hints[key]['min'] = float(text)

    def update_param_hint_max(self, pm, key, text):
        pm.param_hints[key]['max'] = float(text)

    def update_param_hint_vary(self, pm, key, state):
        pm.param_hints[key]['vary'] = state

    def update_param_hint_expr(self, pm, key, text):
        pm.param_hints[key]['expr'] = text

    def clear(self):
        """Clears all data from the main layout."""
        self.clear_layout(self.main_layout)

class DataframeTableWidget(QWidget):
    """Class to display a given dataframe in GUI via QTableWidget.

    This class allows a pandas DataFrame to be displayed within a QTableWidget,
    which is added to a specified layout in your GUI. It provides functionality
    to copy selected data to the clipboard using a context menu or a keyboard shortcut.

    Attributes:
        layout (QVBoxLayout): The layout in the main_app where the QTableWidget will be placed.
    """

    def __init__(self, layout):
        super().__init__()
        self.external_layout = layout
        self.initUI()

    def initUI(self):
        """Initializes the user interface by creating and configuring the QTableWidget."""
        # Clear existing widgets from external layout
        while self.external_layout.count():
            item = self.external_layout.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()

        # Set the internal layout
        layout = QVBoxLayout(self)
        self.setLayout(layout)

        # Create QTableWidget
        self.table_widget = QTableWidget()
        self.table_widget.setSizeAdjustPolicy(QTableWidget.AdjustToContents)
        layout.addWidget(self.table_widget)

        # Enable copy action via context menu
        self.table_widget.setContextMenuPolicy(Qt.ActionsContextMenu)
        copy_action = QAction("Copy", self)
        copy_action.triggered.connect(self.copy_data)
        self.table_widget.addAction(copy_action)

        # Add this widget to the external layout
        self.external_layout.addWidget(self)

    
    def show(self, df, fill_colors=True):
        """Populates the QTableWidget with data from the given DataFrame, with colored columns."""
        if df is not None and not df.empty:
            self.table_widget.setRowCount(df.shape[0])
            self.table_widget.setColumnCount(df.shape[1])
            self.table_widget.setHorizontalHeaderLabels(df.columns)

            # Assign colors based on the column prefix
            column_colors = self.get_column_colors(df.columns)

            for row in range(df.shape[0]):
                for col in range(df.shape[1]):
                    item = QTableWidgetItem(str(df.iat[row, col]))
                    if fill_colors:
                        item.setBackground(column_colors[col])  # Set background color
                    self.table_widget.setItem(row, col, item)

            self.table_widget.resizeColumnsToContents()
        else:
            self.clear()
            
    def get_column_colors(self, columns):
        """Generates a color for each column """
        palette = ['#bda16d', '#a27ba0', '#cb5b12', '#23993b', '#008281', '#147ce4']
        prefix_colors = {}
        column_colors = []

        for col_name in columns:
            prefix = col_name.split("_")[0] if "_" in col_name else col_name
            # Assign the next available color from the palette
            if prefix not in prefix_colors:
                color_index = len(prefix_colors) % len(palette)
                prefix_colors[prefix] = QColor(palette[color_index])
            column_colors.append(prefix_colors[prefix])
        return column_colors

    
    def clear(self):
        """Clears all data from the QTableWidget."""
        self.table_widget.clearContents()
        self.table_widget.setRowCount(0)
        self.table_widget.setColumnCount(0)
        self.table_widget.setHorizontalHeaderLabels([])

    def copy_data(self):
        """Copies selected data from the QTableWidget to the clipboard."""
        selected_indexes = self.table_widget.selectedIndexes()
        if not selected_indexes:
            return

        # Collect unique rows and columns
        rows = set(index.row() for index in selected_indexes)
        cols = set(index.column() for index in selected_indexes)

        data = []
        for row in sorted(rows):
            row_data = []
            for col in sorted(cols):
                item = self.table_widget.item(row, col)
                if item is not None:
                    row_data.append(item.text())
                else:
                    row_data.append('')
            data.append('\t'.join(row_data))

        # Join all rows with newline character and copy to clipboard
        clipboard = QApplication.clipboard()
        clipboard.setText('\n'.join(data))

    def keyPressEvent(self, event):
        """Handles key press events to enable copying with Ctrl+C."""
        if event.modifiers() & Qt.ControlModifier and event.key() == Qt.Key_C:
            self.copy_data()
        else:
            super().keyPressEvent(event)

        
class ColorDelegate(QStyledItemDelegate):
    """Show color in background of color selector comboboxes."""
    def paint(self, painter, option, index):
        painter.save()
        color = index.data(Qt.BackgroundRole)
        if color:
            painter.fillRect(option.rect, color)
        painter.drawText(option.rect, Qt.AlignCenter,
                         index.data(Qt.DisplayRole))
        painter.restore()

    def sizeHint(self, option, index):
        return QSize(70, 20)


class Graph(QWidget):
    """Class to create and handle plot objects.

    This class provides functionality to create and customize plots using
    matplotlib
    and seaborn libraries within a Pyside6-based GUI application. It supports
    plotting
    various styles such as point plots, scatter plots, box plots, line plots,
    bar plots,
    trendline plots, and wafer plots.
    The class allows customization of plot properties including titles,
    labels, axis limits, grid display, legend appearance, color
    palettes, and more. It also supports multiple y-axis plotting and the
    option to show
    trendline equations.
    """

    def __init__(self, graph_id=None):
        super().__init__()
        self.df_name = None
        self.filters = {}  # List of filter
        self.graph_id = graph_id
        self.plot_width = 600
        self.plot_height = 500
        self.plot_style = "point"
        self.x = None
        self.y = []  # Multiple y column allowing to plot multiples lines
        self.z = None
        self.xmin = None
        self.xmax = None
        self.ymin = None
        self.ymax = None
        self.zmin = None
        self.zmax = None

        self.plot_title = None
        self.xlabel = None
        self.ylabel = None
        self.zlabel = None

        self.y2 = None  # Secondary y-axis
        self.y3 = None  # Tertiary y-axis
        self.y2min = None
        self.y2max = None
        self.y3min = None
        self.y3max = None
        self.y2label = None  # Secondary y-axis
        self.y3label = None  # Tertiary y-axis

        self.x_rot = 0
        self.grid = False
        self.legend_visible = True
        self.legend_location = 'upper right'
        self.legend_outside = False
        self.legend_properties = []

        self.color_palette = "jet"  # Palette for wafer maps
        self.dpi = 100
        self.wafer_size = 300
        self.wafer_stats = True
        self.trendline_order = 1
        self.show_trendline_eq = True
        self.show_bar_plot_error_bar = True
        self.join_for_point_plot = False

        self.figure = None
        self.ax = None
        self.ax2 = None  # Secondary y-axis
        self.ax3 = None  # Tertiary y-axis
        self.canvas = None
        self.graph_layout = QVBoxLayout()  # Layout for store plot
        self.setLayout(self.graph_layout)

        # Set layout margins to 0 to remove extra space
        self.graph_layout.setContentsMargins(0, 0, 0, 0)
        self.graph_layout.setSpacing(0)
        
    def save(self, fname=None):
        """ Save Graph object to serialization. Save it if a fname is given """
        # List of keys to exclude from serialization
        excluded_keys = ['figure', 'canvas', 'setLayout', 'graph_layout',
                         'some_signal_instance']

        dict_graph = {}
        for key, val in vars(self).items():
            if key not in excluded_keys and not callable(val):
                try:
                    json.dumps(val)
                    dict_graph[key] = val
                except TypeError:
                    continue

        if fname is not None:
            with open(fname, 'w') as f:
                json.dump(dict_graph, f, indent=4)

        return dict_graph

    def set_attributes(self, attributes_dict):
        """Set attributes of the Graph object from a given dictionary."""
        
        for key, value in attributes_dict.items():
            if hasattr(self, key):
                setattr(self, key, value)

    def clear_layout(self, layout):
        """Clears all widgets and layouts from the specified layout."""
        if layout is not None:
            while layout.count():
                item = layout.takeAt(0)
                widget = item.widget()
                if widget:
                    widget.deleteLater()
                else:
                    self.clear_layout(item.layout())

    def create_plot_widget(self, dpi, layout=None):
        """Creates a new plot canvas with the specified DPI and adds it to
        the specified layout or the default graph_layout"""
        if dpi:
            self.dpi = dpi
        else:
            self.dpi = 100
        self.clear_layout(self.graph_layout)
        plt.close('all')

        self.figure = plt.figure(dpi=self.dpi)
        self.ax = self.figure.add_subplot(111)
        self.canvas = FigureCanvas(self.figure)
        self.figure.subplots_adjust(left=0, right=1, top=1, bottom=0)

        self.toolbar = NavigationToolbar2QT(self.canvas, self)
        for action in self.toolbar.actions():
            if action.text() in ['Save', 'Subplots', 'Customize']:
                action.setVisible(False)

        if layout:
            layout.addWidget(self.canvas)
            layout.addWidget(self.toolbar)
        else:
            self.graph_layout.addWidget(self.canvas)
            self.graph_layout.addWidget(self.toolbar)

        self.canvas.figure.tight_layout()
        self.canvas.draw_idle()

    def plot(self, df):
        """Updates the plot based on the provided DataFrame and plot
        settings."""
        self.ax.clear()
        if self.ax2:
            self.ax2.clear()
        if self.ax3:
            self.ax3.clear()

        if self.df_name is not None and self.x is not None and self.y is not \
                None:
            self._plot_primary_axis(df)
            self._plot_secondary_axis(df)
            self._plot_tertiary_axis(df)
        else:
            self.ax.plot([], [])

        self._set_limits()
        self._set_labels()
        self._set_grid()
        self._set_rotation()
        self._set_legend()

        self.get_legend_properties()
        self.ax.get_figure().tight_layout()
        self.canvas.draw_idle()

    def get_legend_properties(self):
        """Retrieves properties of each existing legend item."""
        legend_properties = []
        if self.ax:
            legend = self.ax.get_legend()
            if legend:
                legend_texts = legend.get_texts()
                legend_handles = legend.legendHandles
                for idx, text in enumerate(legend_texts):

                    label = text.get_text()
                    handle = legend_handles[idx]
                    if self.plot_style in ['point', 'scatter', 'line']:
                        color = handle.get_markerfacecolor()
                        marker = handle.get_marker()
                    # Box & bar plots do not use markers → set defautl values
                    elif self.plot_style in ['box', 'bar']:
                        color = rgba_to_named_color(handle.get_facecolor())
                        marker = 'o'
                    else:
                        color = 'blue'
                        marker = 'o'
                    legend_properties.append(
                        {'label': label, 'marker': marker, 'color': color})
        self.legend_properties = legend_properties
        return self.legend_properties

    def customize_legend_via_gui(self, main_layout):
        """Displays legend properties in the GUI for user modifications."""
        self.clear_layout(main_layout)
        headers = ['Label', 'Marker', 'Color']
        # Create vertical layouts for each property type
        label_layout = QVBoxLayout()
        marker_layout = QVBoxLayout()
        color_layout = QVBoxLayout()
        for header in headers:
            label = QLabel(header)
            label.setAlignment(Qt.AlignCenter)
            if header == "Label":
                label_layout.addWidget(label)
            elif header == "Marker":
                if self.plot_style == 'point':
                    marker_layout.addWidget(label)
                else:
                    pass
            elif header == "Color":
                color_layout.addWidget(label)

        for idx, prop in enumerate(self.legend_properties):
            # LABEL
            label = QLineEdit(prop['label'])
            label.setFixedWidth(200)
            label.textChanged.connect(
                lambda text, idx=idx: self.udp_legend(idx, 'label', text))
            label_layout.addWidget(label)

            if self.plot_style == 'point':
                # MARKER
                marker = QComboBox()
                marker.addItems(MARKERS)  # Add more markers as needed
                marker.setCurrentText(prop['marker'])
                marker.currentTextChanged.connect(
                    lambda text, idx=idx: self.udp_legend(idx, 'marker', text))
                marker_layout.addWidget(marker)
            else:
                pass

            # COLOR
            color = QComboBox()
            delegate = ColorDelegate(color)
            color.setItemDelegate(delegate)
            for color_code in DEFAULT_COLORS:
                item = color.addItem(color_code)
                item = color.model().item(color.count() - 1)
                item.setBackground(QColor(color_code))

            color.setCurrentText(prop['color'])
            color.currentIndexChanged.connect(
                lambda idx, color=color: self.update_combobox_color(color))

            color.currentTextChanged.connect(
                lambda text, idx=idx: self.udp_legend(idx, 'color', text))
            color_layout.addWidget(color)

            # Ensure the color is updated on load
            self.update_combobox_color(color)

        # Add vertical layouts to main layout
        main_layout.addLayout(label_layout)
        main_layout.addLayout(marker_layout)
        main_layout.addLayout(color_layout)

    def update_combobox_color(self, combobox):
        """Update combobox background color based on the selected color."""
        selected_color = combobox.currentText()
        color = QColor(selected_color)
        palette = combobox.palette()
        palette.setColor(QPalette.Button, color)
        palette.setColor(QPalette.ButtonText, Qt.white)
        combobox.setAutoFillBackground(True)
        combobox.setPalette(palette)
        combobox.update()

    def udp_legend(self, idx, property_type, text):
        """Updates legend properties based on user modifications via GUI."""
        self.legend_properties[idx][property_type] = text
        self._set_legend()

    def _plot_primary_axis(self, df):
        """Plots data on the primary axis based on the current plot style."""
        if not self.legend_properties:
            markers = DEFAULT_MARKERS
            colors = DEFAULT_COLORS
        else:
            markers = [str(prop['marker']) for prop in self.legend_properties]
            colors = [str(prop['color']) for prop in self.legend_properties]
        for y in self.y:
            if self.plot_style == 'point':
                sns.pointplot(data=df, x=self.x, y=y, hue=self.z,
                              ax=self.ax,
                              linestyles='-' if self.join_for_point_plot
                              else 'none', marker=markers, palette=colors,
                              markeredgecolor='black', markeredgewidth=1,
                              err_kws={'linewidth': 1, 'color': 'black'},
                              capsize=0.02)
            elif self.plot_style == 'scatter':
                sns.scatterplot(data=df, x=self.x, y=y, hue=self.z, ax=self.ax,
                                s=100, edgecolor='black', palette=colors)
            elif self.plot_style == 'box':
                sns.boxplot(data=df, x=self.x, y=y, hue=self.z, 
                            ax=self.ax, palette=colors, width=0.4)
            elif self.plot_style == 'line':
                sns.lineplot(data=df, x=self.x, y=y, hue=self.z, ax=self.ax,
                             palette=colors)
            elif self.plot_style == 'bar':
                sns.barplot(data=df, x=self.x, y=y, hue=self.z,
                            errorbar='sd' if self.show_bar_plot_error_bar
                            else None, err_kws={'linewidth': 1},
                            ax=self.ax, palette=colors)

            elif self.plot_style == 'trendline':
                sns.regplot(data=df, x=self.x, y=y, ax=self.ax, scatter=True,
                            order=self.trendline_order)
                if self.show_trendline_eq:
                    self._annotate_trendline_eq(df)
            elif self.plot_style == 'wafer':
                self._plot_wafer(df)
                
            elif self.plot_style == '2Dmap':
                x_col = self.x 
                y_col = y if isinstance(self.y, list) else self.y  
                z_col = self.z 
                xmin = df[x_col].min()
                xmax = df[x_col].max()
                ymin = df[y_col].min()
                ymax = df[y_col].max()
                heatmap_data = df.pivot(index=y_col, columns=x_col, values=z_col)
                vmin = self.zmin if self.zmin else heatmap_data.min().min()
                vmax = self.zmax if self.zmax else heatmap_data.max().max()
                
                heatmap = self.ax.imshow(heatmap_data, aspect='equal', extent=[xmin, xmax, ymin, ymax], cmap=self.color_palette, origin='lower', vmin=vmin, vmax=vmax)
                plt.colorbar(heatmap, orientation='vertical')
            else:
                show_alert("Unsupported plot style")

    def _set_legend(self):
        """Sets up and displays the legend for the plot."""
        handles, labels = self.ax.get_legend_handles_labels()
        if self.ax2:
            handles2, labels2 = self.ax2.get_legend_handles_labels()
            handles += handles2
            labels += labels2
            self.ax2.legend().remove()  # Turn off legend for ax2
        if self.ax3:
            handles3, labels3 = self.ax3.get_legend_handles_labels()
            handles += handles3
            labels += labels3
            self.ax3.legend().remove()  # Turn off legend for ax3
        if handles:
            legend_labels = []
            if self.legend_properties:
                try:
                    for idx, prop in enumerate(self.legend_properties):
                        legend_labels.append(prop['label'])
                        handles[idx].set_label(
                            prop['label'])  # Set legend label
                        handles[idx].set_color(prop['color'])  # Set color
                        if self.plot_style in ['point', 'scatter']:
                            handles[idx].set_marker(
                                prop['marker'])  # Set marker
                        else:
                            pass
                except Exception as e:
                    self.legend_properties = []
                    legend_labels = labels
                    self.legend_properties = self.get_legend_properties()

            else:
                legend_labels = labels
                self.legend_properties = self.get_legend_properties()

            if self.legend_visible:
                self.ax.legend(handles, legend_labels, loc=self.legend_location)
            else:
                self.ax.legend().remove()
            if self.legend_outside:
                self.ax.legend(handles, legend_labels, loc='center left',
                               bbox_to_anchor=(1, 0.5))

    def _set_grid(self):
        """Add grid for the plot"""
        if self.grid:
            self.ax.grid(True, linestyle='--', linewidth=0.5, color='gray')
        else:
            self.ax.grid(False)

    def _set_rotation(self):
        """Set rotation of the ticklabels of the x axis"""
        if self.x_rot != 0:
            plt.setp(self.ax.get_xticklabels(), rotation=self.x_rot, ha="right",
                     rotation_mode="anchor")

    def _annotate_trendline_eq(self, df):
        """Add the trendline equation in the plot"""
        x_data = df[self.x]
        y_data = df[self.y[0]]
        coefficients = np.polyfit(x_data, y_data, self.trendline_order)
        equation = 'y = '
        for i, coeff in enumerate(coefficients[::-1]):
            equation += (f'{coeff:.2f}x^{self.trendline_order - i} + '
                         if i < self.trendline_order else f'{coeff:.2f}')
        self.ax.annotate(equation, xy=(0.02, 0.95), xycoords='axes fraction',
                         fontsize=10, color='blue')

    def _plot_wafer(self, df):
        """PLot wafer plot by creating an object of WaferPlot Class"""
        vmin = self.zmin if self.zmin else None
        vmax = self.zmax if self.zmax else None
        wdf = WaferPlot()
        wdf.plot(self.ax, x=df[self.x], y=df[self.y[0]], z=df[self.z],
                 cmap=self.color_palette,
                 vmin=vmin, vmax=vmax, stats=self.wafer_stats,
                 r=(self.wafer_size / 2))

    def _set_limits(self):
        """Set the limits of axis"""
        if self.xmin and self.xmax:
            self.ax.set_xlim(float(self.xmin), float(self.xmax))
        if self.ymin and self.ymax:
            self.ax.set_ylim(float(self.ymin), float(self.ymax))
        if self.ax2 and self.y2min and self.y2max:
            self.ax2.set_ylim(float(self.y2min), float(self.y2max))
        if self.ax3 and self.y3min and self.y3max:
            self.ax3.set_ylim(float(self.y3min), float(self.y3max))

    def _set_labels(self):
        """Set titles and labels for axis and plot"""
        if self.plot_style == 'wafer':
            if self.plot_title:
                self.ax.set_title(self.plot_title)
            else:
                self.ax.set_title(self.z)
                
        elif self.plot_style == '2Dmap':
            if self.plot_title:
                self.ax.set_title(self.plot_title)
            else:
                self.ax.set_title(self.z) 
            if self.xlabel:
                self.ax.set_xlabel(self.xlabel)
            else:
                self.ax.set_xlabel(self.x)
            if self.ylabel:
                self.ax.set_ylabel(self.ylabel)
            else:
                self.ax.set_ylabel(self.y[0])
        else:
            self.ax.set_title(self.plot_title)
            if self.xlabel:
                self.ax.set_xlabel(self.xlabel)
            else:
                self.ax.set_xlabel(self.x)
            if self.ylabel:
                self.ax.set_ylabel(self.ylabel)
            else:
                self.ax.set_ylabel(self.y[0])

    def _plot_secondary_axis(self, df):
        if self.ax2:
            self.ax2.remove()
            self.ax2 = None
        if hasattr(self, 'y2') and self.y2:
            self.ax2 = self.ax.twinx()
            if self.plot_style == 'line':
                sns.lineplot(data=df, x=self.x, y=self.y2, hue=self.z,
                             ax=self.ax2, color='red')
            elif self.plot_style == 'point':
                sns.pointplot(
                    data=df, x=self.x, y=self.y2, hue=self.z, ax=self.ax2,
                    linestyles='-' if self.join_for_point_plot else 'none',
                    marker='s', color='red', markeredgecolor='black',
                    markeredgewidth=1,
                    dodge=True, err_kws={'linewidth': 1, 'color': 'black'},
                    capsize=0.02
                )
            elif self.plot_style == 'scatter':
                sns.scatterplot(data=df, x=self.x, y=self.y2, hue=self.z,
                                ax=self.ax2,
                                s=100, edgecolor='black', color='red')
            else:
                self.ax2.remove()
                self.ax2 = None

            self.ax2.set_ylabel(self.y2label, color='red')
            self.ax2.tick_params(axis='y', colors='red')

    def _plot_tertiary_axis(self, df):
        if self.ax3:
            self.ax3.remove()
            self.ax3 = None
        if hasattr(self, 'y3') and self.y3:
            self.ax3 = self.ax.twinx()
            self.ax3.spines["right"].set_position(("outward", 100))
            if self.plot_style == 'line':
                sns.lineplot(data=df, x=self.x, y=self.y3, hue=self.z,
                             ax=self.ax3, color='green')
            elif self.plot_style == 'point':
                sns.pointplot(
                    data=df, x=self.x, y=self.y3, hue=self.z, ax=self.ax3,
                    linestyles='-' if self.join_for_point_plot else 'none',
                    marker='s', color='red', markeredgecolor='black',
                    markeredgewidth=1,
                    dodge=True, err_kws={'linewidth': 1, 'color': 'black'},
                    capsize=0.02
                )
            elif self.plot_style == 'scatter':
                sns.scatterplot(data=df, x=self.x, y=self.y3, hue=self.z,
                                ax=self.ax3,
                                s=100, edgecolor='black', color='green')
            else:
                self.ax3.remove()
                self.ax3 = None
            self.ax3.set_ylabel(self.y3label, color='green')
            self.ax3.tick_params(axis='y', colors='green')

    

class FitModelManager:
    """
    Class to manage fit models created by USERS.

    Attributes:
    settings (QSettings): QSettings object to store and retrieve settings.
    default_model_folder (str): Default folder path where fit models are stored.
    available_models (list): List of available fit model filenames in the
    default folder.
    """

    def __init__(self, settings):
        self.settings = settings
        self.default_model_folder = self.settings.value("default_model_folder","")
        self.available_models = []
        if self.default_model_folder:
            self.scan_models()

    def set_default_model_folder(self, folder_path):
        """
        Set the default folder path where fit models will be stored.

        Args:
        folder_path (str): Path to the default folder.
        """
        self.default_model_folder = folder_path
        self.settings.setValue("default_model_folder", folder_path)
        self.scan_models()

    def scan_models(self):
        """
        Scan the default folder and populate the available_models list.

        This method scans the default_model_folder for files with the '.json'
        extension and updates the available_models list accordingly.
        """
        self.available_models = []

        if self.default_model_folder:
            if not os.path.exists(self.default_model_folder):
                # Folder is specified but does not exist anymore (deleted or renamed)
                msg= f"Default 'Fit_models' folder '{self.default_model_folder}' not found. Please specify another one in the 'More Settings' tab."
                show_alert(msg)
                # Reset the default model folder to empty
                self.default_model_folder = ""
                self.settings.setValue("default_model_folder", "")
                return  # Exit the method since the folder is missing
            
            # Scan the folder for JSON files if it exists
            try:
                for file_name in os.listdir(self.default_model_folder):
                    if file_name.endswith('.json'):
                        self.available_models.append(file_name)
            except Exception as e:
                print(f"Error scanning the folder '{self.default_model_folder}': {e}")


    def get_available_models(self):
        """
        Retrieve the list of available fit model filenames.

        Returns:
        list: List of available fit model filenames in the default folder.
        """
        return self.available_models


class CommonUtilities():
    """ Class contain all common methods or utility codes used other modules"""
    
    
    
    def reinit_spectrum(self, fnames, spectrums):
        """Reinitilize a FITSPY spectrum object"""
        for fname in fnames:
            spectrum, _ = spectrums.get_objects(fname)
            spectrum.reinit()
            spectrum.baseline.mode = "Linear"

    def clear_layout(self, layout):
        """Clear everything in a given Qlayout"""
        if layout is not None:
            for i in reversed(range(layout.count())):
                item = layout.itemAt(i)
                if isinstance(item.widget(),
                              (FigureCanvas, NavigationToolbar2QT)):
                    widget = item.widget()
                    layout.removeWidget(widget)
                    widget.close()

    def replace_peak_labels(self, fit_model, param):
        """Replace prefix 'm01' of peak model by labels designed by user"""
        peak_labels = fit_model["peak_labels"]
        if "_" in param:
            prefix, param = param.split("_", 1)  
            # Convert prefix to peak_label
            peak_index = int(prefix[1:]) - 1
            if 0 <= peak_index < len(peak_labels):
                peak_label = peak_labels[peak_index]
                return f"{param}_{peak_label}"
        return param

    def quadrant(self, row):
        """Define 4 quadrant of a wafer"""
        if row['X'] < 0 and row['Y'] < 0:
            return 'Q1'
        elif row['X'] < 0 and row['Y'] > 0:
            return 'Q2'
        elif row['X'] > 0 and row['Y'] > 0:
            return 'Q3'
        elif row['X'] > 0 and row['Y'] < 0:
            return 'Q4'
        else:
            return np.nan

    def zone(self, row, radius):
        """Define 3 zones (Center, Mid-Radius, Edge) based on X and Y
        coordinates."""
        r = radius
        x = row['X']
        y = row['Y']
        distance_to_center = np.sqrt(x ** 2 + y ** 2)
        if distance_to_center <= r * 0.35:
            return 'Center'
        elif distance_to_center > r * 0.35 and distance_to_center < r * 0.8:
            return 'Mid-Radius'
        elif distance_to_center >= 0.8 * r:
            return 'Edge'
        else:
            return np.nan

    def display_df_in_table(self, table_widget, df_results):
        """Display pandas DataFrame in QTableWidget in GUI"""
        table_widget.setRowCount(df_results.shape[0])
        table_widget.setColumnCount(df_results.shape[1])
        table_widget.setHorizontalHeaderLabels(df_results.columns)
        for row in range(df_results.shape[0]):
            for col in range(df_results.shape[1]):
                item = QTableWidgetItem(str(df_results.iat[row, col]))
                table_widget.setItem(row, col, item)
        table_widget.resizeColumnsToContents()

    def view_text(self, ui, title, text):
        """ Create a QTextBrowser to display a text content"""
        report_viewer = QDialog(ui)
        report_viewer.setWindowTitle(title)
        report_viewer.setGeometry(100, 100, 800, 600)
        text_browser = QTextBrowser(report_viewer)
        text_browser.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        text_browser.setOpenExternalLinks(True)
        text_browser.setPlainText(text)
        text_browser.moveCursor(QTextCursor.Start)
        layout = QVBoxLayout(report_viewer)
        layout.addWidget(text_browser)
        report_viewer.show()

    def view_markdown(self, ui, title, fname, x, y, working_folder):
        """To convert MD file to html format and display them in GUI"""
        with open(fname, 'r', encoding='utf-8') as f:
            markdown_content = f.read()
        html_content = markdown.markdown(markdown_content)
        DIRNAME = os.path.dirname(__file__)
        html_content = html_content.replace('src="',
                                            f'src="'
                                            f'{os.path.join(DIRNAME, working_folder)}')
        about_dialog = QDialog(ui)
        about_dialog.setWindowTitle(title)
        about_dialog.resize(x, y)
        text_browser = QTextBrowser(about_dialog)
        text_browser.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        text_browser.setOpenExternalLinks(True)
        text_browser.setHtml(html_content)
        layout = QVBoxLayout(about_dialog)
        layout.addWidget(text_browser)
        about_dialog.setLayout(layout)
        about_dialog.show()

    def dark_palette(self):
        """Palette color for dark mode of the appli's GUI"""
        dark_palette = QPalette()
        dark_palette.setColor(QPalette.Window, QColor(70, 70, 70))
        dark_palette.setColor(QPalette.WindowText, Qt.white)
        dark_palette.setColor(QPalette.Base,
                              QColor(65, 65, 65))  # QlineEdit Listbox bg
        dark_palette.setColor(QPalette.AlternateBase, QColor(45, 45, 45))
        dark_palette.setColor(QPalette.ToolTipBase, QColor(255, 255, 220))
        dark_palette.setColor(QPalette.ToolTipText, Qt.white)
        dark_palette.setColor(QPalette.Text, Qt.white)
        dark_palette.setColor(QPalette.Button, QColor(64, 64, 64))
        dark_palette.setColor(QPalette.ButtonText, Qt.white)
        dark_palette.setColor(QPalette.BrightText, Qt.red)
        dark_palette.setColor(QPalette.Link, QColor(42, 130, 218))
        dark_palette.setColor(QPalette.Highlight, QColor(42, 130, 218))
        dark_palette.setColor(QPalette.HighlightedText, Qt.white)
        dark_palette.setColor(QPalette.PlaceholderText, QColor(140, 140, 140))
        dark_palette.setColor(QPalette.Base, QColor(60, 60, 60))  # Background color for QMenu
        
        return dark_palette

    def light_palette(self):
        """Palette color for light mode of the appli's GUI"""
        light_palette = QPalette()
        light_palette.setColor(QPalette.Window, QColor(225, 225, 225))
        light_palette.setColor(QPalette.WindowText, Qt.black)
        light_palette.setColor(QPalette.Base, QColor(215, 215, 215))
        light_palette.setColor(QPalette.AlternateBase, QColor(230, 230, 230))
        light_palette.setColor(QPalette.ToolTipBase, QColor(255, 255, 220))
        light_palette.setColor(QPalette.ToolTipText, Qt.black)
        light_palette.setColor(QPalette.Text, Qt.black)
        light_palette.setColor(QPalette.Button, QColor(230, 230, 230))
        light_palette.setColor(QPalette.ButtonText, Qt.black)
        light_palette.setColor(QPalette.BrightText, Qt.red)
        light_palette.setColor(QPalette.Link, QColor(42, 130, 218))
        light_palette.setColor(QPalette.Highlight, QColor(42, 130, 218))
        light_palette.setColor(QPalette.HighlightedText, Qt.black)
        light_palette.setColor(QPalette.PlaceholderText, QColor(150, 150, 150))
        light_palette.setColor(QPalette.Base, QColor(240, 240, 240))  # Menu background color

        return light_palette


class FitThread(QThread):
    """ Class to perform fitting in a separate Thread """
    progress_changed = Signal(int)

    def __init__(self, spectrums, fit_model, fnames, ncpus=1):
        super().__init__()
        self.spectrums = spectrums
        self.fit_model = fit_model
        self.fnames = fnames
        self.ncpus = ncpus

    def run(self):
        fit_model = deepcopy(self.fit_model)
        self.spectrums.apply_model(fit_model, fnames=self.fnames,
                                   ncpus=self.ncpus, show_progressbar=False)

        self.progress_changed.emit(100)


class WaferPlot:
    """Class to plot wafer map"""

    def __init__(self, inter_method='linear'):
        self.inter_method = inter_method  # Interpolation method

    def plot(self, ax, x, y, z, cmap="jet", r=100, vmax=None, vmin=None,
             stats=True):
        """
        Plot a wafer map on the provided axes.

        Args:
        ax (matplotlib.axes.Axes): Axes object to plot the wafer map.
        x (array-like): X-coordinates of measurement points.
        y (array-like): Y-coordinates of measurement points.
        z (array-like): measurement data corresponding to (x,y) points.
        cmap (str, optional): Colormap for the plot. Defaults to "jet".
        r (float, optional): Radius of the wafer in millimeters.
        vmax (float, optional): Maximum value for the color scale.
        vmin (float, optional): Minimum value for the color scale.
        stats (bool, optional): Display statistical values on the plot.
        """
        # Generate a meshgrid for the wafer and Interpolate z onto the meshgrid
        xi, yi = np.meshgrid(np.linspace(-r, r, 300), np.linspace(-r, r, 300))
        zi = griddata((x, y), z, (xi, yi), method=self.inter_method)

        # Plot the wafer map
        im = ax.imshow(zi, extent=[-r - 1, r + 1, -r - 0.5, r + 0.5],
                       origin='lower', cmap=cmap, interpolation='nearest')

        # Add open circles corresponding to measurement points
        ax.scatter(x, y, facecolors='none', edgecolors='black', s=20)

        # Add a circle as a decoration
        wafer_circle = patches.Circle((0, 0), radius=r, fill=False,
                                      color='black', linewidth=1)
        ax.add_patch(wafer_circle)

        ax.set_ylabel("Wafer size (mm)")

        # Remove unnecessary axes
        ax.spines['right'].set_visible(False)
        ax.spines['top'].set_visible(False)
        ax.spines['bottom'].set_visible(False)
        ax.tick_params(axis='x', which='both', bottom=False, top=False)
        ax.tick_params(axis='y', which='both', right=False, left=True)
        ax.set_xticklabels([])

        # Set color scale limits if provided
        if vmax is not None and vmin is not None:
            im.set_clim(vmin, vmax)
        plt.colorbar(im, ax=ax)
        if stats:
            self.stats(z, ax)

    def stats(self, z, ax):
        """
        Calculate and display statistical values in the wafer plot.
        """
        # Calculate statistical values
        mean_value = z.mean()
        max_value = z.max()
        min_value = z.min()
        sigma_value = z.std()
        three_sigma_value = 3 * sigma_value

        # Annotate the plot with statistical values
        ax.text(0.05, -0.1, f"3\u03C3: {three_sigma_value:.2f}",
                transform=ax.transAxes,
                fontsize=10, verticalalignment='bottom')
        ax.text(0.99, -0.1, f"Max: {max_value:.2f}",
                transform=ax.transAxes,
                fontsize=10, verticalalignment='bottom',
                horizontalalignment='right')
        ax.text(0.99, -0.05, f"Min: {min_value:.2f}",
                transform=ax.transAxes,
                fontsize=10, verticalalignment='bottom',
                horizontalalignment='right')
        ax.text(0.05, -0.05, f"Mean: {mean_value:.2f}",
                transform=ax.transAxes, fontsize=10, verticalalignment='bottom')


class CustomListWidget(QListWidget):
    """
    Customized QListWidget with drag-and-drop functionality for rearranging
    items.

    This class inherits from QListWidget and provides extended functionality
    for reordering items via drag-and-drop operations.

    Signals:
        items_reordered:
            Emitted when items in the list widget are reordered by the user
            using drag-and-drop.
    """
    items_reordered = Signal()
    files_dropped = Signal(list) 

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)  # Enable external drag-drop
        self.setDragDropMode(QListWidget.InternalMove)
        self.setSelectionMode(QAbstractItemView.ExtendedSelection)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            super().dragEnterEvent(event)
            
    def dragMoveEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            super().dragMoveEvent(event)
            
    def dropEvent(self, event):
        """
        Overrides the dropEvent method to emit the items_reordered signal
            after an item is dropped into a new position.
        """
        if event.mimeData().hasUrls():
            file_paths = [url.toLocalFile() for url in event.mimeData().urls()]
            self.files_dropped.emit(file_paths)  # emit signal with file list
            event.acceptProposedAction()
        else:
            super().dropEvent(event)
            self.items_reordered.emit()
    
class ConvertFile():
    def __init__(self, ui, settings):
        super().__init__()
        self.ui = ui
        self.settings = settings
        self.selected_file_paths = []
        
        self.ui.btn_browse.clicked.connect(lambda: self.browse_files())
        self.ui.btn_convert.clicked.connect(lambda: self.convert())

        
    def browse_files(self, file_paths=None):
        """
        Browse and select files to convert. Populates listbox with file names.
        """
        # Ensure file_paths is None or a list
        if file_paths is not None and not isinstance(file_paths, list):
            print(f"Warning: Invalid file_paths type: {type(file_paths)}. Expected list or None.")
            return

        self.selected_file_paths.clear()
        self.ui.listbox.clear()

        # Show dialog if no file_paths provided
        if file_paths is None:
            last_dir = self.settings.value("last_directory", "/")
            options = QFileDialog.Options()
            options |= QFileDialog.ReadOnly
            file_paths, _ = QFileDialog.getOpenFileNames(
                self.ui.tabWidget,
                "Open 2D Map File(s) to Convert",
                last_dir,
                "Text Files (*.txt)",
                options=options
            )

        # If files were selected
        if file_paths:
            last_dir = QFileInfo(file_paths[0]).absolutePath()
            self.settings.setValue("last_directory", last_dir)
            
            for file_path in file_paths:
                if os.path.isfile(file_path):
                    abs_path = os.path.abspath(file_path)
                    self.selected_file_paths.append(abs_path)
                    self.ui.listbox.addItem(os.path.basename(abs_path))


    def convert(self, output_dir=None):
        """
        Convert selected files and add output file names to listbox.
        """
        if not self.selected_file_paths:
            QMessageBox.warning(self.ui.tabWidget, "No Files", "Please select files first.")
            return

        for input_path in self.selected_file_paths:
            selected_file_name = os.path.basename(input_path)
            if selected_file_name.endswith('_converted.txt'):
                continue

            name, _ = os.path.splitext(selected_file_name)
            if output_dir is None:
                output_dir = os.path.dirname(input_path)
            output_path = os.path.join(output_dir, f"{name}_converted.txt")

            with tempfile.TemporaryDirectory() as dirtemp:
                temp_input_path = os.path.join(dirtemp, selected_file_name)
                shutil.copy(input_path, temp_input_path)

                # Clean tab formatting
                with open(temp_input_path, 'r+', encoding='utf-8') as fid:
                    data = fid.read().replace("\t\t", "\t")
                    fid.seek(0)
                    fid.write(data)
                    fid.truncate()

                try:
                    self.convert_action(temp_input_path, output_path)
                    self.ui.listbox.addItem(os.path.basename(output_path))
                except Exception as e:
                    QMessageBox.critical(self.ui.tabWidget, "Conversion Error", f"Failed to convert {selected_file_name}\n{str(e)}")

    def convert_action(self, input_path, output_path):
        """
        Converts the file by reshaping intensity values into a matrix
        and saving the output as a reformatted text file.
        """
        try:
            dfr = pd.read_csv(input_path, delimiter="\t")
            assert {'#X', '#Y', '#Wave', '#Intensity'}.issubset(dfr.columns)
        except Exception as e:
            raise ValueError(f"Invalid file format: {str(e)}")

        # Count number of points per spectrum
        dfr_spectrum = dfr.groupby(['#Y', '#X']).size().reset_index(name='size')
        wavenumber_range = dfr_spectrum['size'][1]
        total_spectra_number = len(dfr_spectrum)

        dfr_wavenumbers = dfr['#Wave'][:wavenumber_range]
        dfr_intensity = dfr['#Intensity']

        # Build intensity matrix
        intensity_data = {
            str(dfr_wavenumbers.iloc[i]): [
                dfr_intensity[i + j * wavenumber_range] for j in range(total_spectra_number)
            ]
            for i in range(wavenumber_range)
        }

        # Create intensity DataFrame and concat with coordinates
        dfr_intensity_matrix = pd.DataFrame(intensity_data)
        dfr_spectrum = pd.concat([dfr_spectrum[['#Y', '#X']].reset_index(drop=True), dfr_intensity_matrix], axis=1)

        # Cleanup
        dfr_spectrum.rename(columns={'#X': '', '#Y': ''}, inplace=True)

        # Export
        dfr_spectrum.to_csv(output_path, sep='\t', encoding='utf-8', index=False)



class CustomizedPalette(QComboBox):
    """Custom QComboBox to show color palette previews along with their names."""

    def __init__(self, palette_list=None, parent=None, icon_size=(99, 12)):
        super().__init__(parent)
        self.icon_width, self.icon_height = icon_size
        self.setIconSize(QSize(*icon_size))
        self.setMinimumWidth(100)

        self.palette_list = palette_list or PALETTE
        self._populate_with_previews()

    def _populate_with_previews(self):
        self.clear()
        for cmap_name in self.palette_list:
            icon = QIcon(self._create_colormap_preview(cmap_name))
            self.addItem(icon, cmap_name)

    def _create_colormap_preview(self, cmap_name):
        """Generate a horizontal gradient preview image for the colormap."""
        width, height = self.icon_width, self.icon_height
        gradient = np.linspace(0, 1, 20).reshape(1, -1)

        fig = Figure(figsize=(width / 100, height / 100), dpi=100)
        canvas = FigureCanvas(fig)
        ax = fig.add_axes([0, 0, 1, 1], frameon=False)
        ax.imshow(gradient, aspect='auto', cmap=cm.get_cmap(cmap_name))
        ax.set_axis_off()
        canvas.draw()

        image = np.array(canvas.buffer_rgba())
        qimage = QImage(image.data, image.shape[1], image.shape[0],
                        QImage.Format_RGBA8888)
        return QPixmap.fromImage(qimage)

    def get_selected_palette(self):
        return self.currentText()

def compress(array):
    """Compress and encode a numpy array to a base64 string."""
    compressed = zlib.compress(array.tobytes())
    encoded = base64.b64encode(compressed).decode('utf-8')
    return encoded


def decompress(data, dtype):
    """Decode and decompress a base64 string to a numpy array."""
    decoded = base64.b64decode(data.encode('utf-8'))
    decompressed = zlib.decompress(decoded)
    return np.frombuffer(decompressed, dtype=dtype)

def plot_baseline_dynamically(ax, spectrum):
    """Evaluate and plot baseline points and line dynamically"""
    if not spectrum.baseline.is_subtracted:
        x_bl = spectrum.x
        y_bl = spectrum.y if spectrum.baseline.attached else None
        if len(spectrum.baseline.points[0]) == 0:
            return
        # Clear any existing baseline plot
        for line in ax.lines:
            if line.get_label() == "Baseline":
                line.remove()
        # Evaluate the baseline
        attached = spectrum.baseline.attached
        baseline_values = spectrum.baseline.eval(x_bl, y_bl,
                                                    attached=attached)
        ax.plot(x_bl, baseline_values, 'r')

        # Plot the attached baseline points
        if spectrum.baseline.attached and y_bl is not None:
            attached_points = spectrum.baseline.attached_points(x_bl, y_bl)
            ax.plot(attached_points[0], attached_points[1], 'ko',
                    mfc='none')
        else:
            ax.plot(spectrum.baseline.points[0],
                    spectrum.baseline.points[1], 'ko', mfc='none', ms=5)

def populate_spectrum_listbox(spectrum, spectrum_name, checked_states):
    """ Populate the listbox with spectrums with colors"""
    item = QListWidgetItem(spectrum_name)            
    item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
    item.setCheckState(checked_states.get(spectrum_name, Qt.Checked))

    if spectrum.baseline.is_subtracted:
        if not hasattr(spectrum.result_fit, 'success'):
            item.setBackground(QColor("red"))
        elif spectrum.result_fit.success:
            item.setBackground(QColor("green"))
        else:
            item.setBackground(QColor("orange"))
    else:
        item.setBackground(QColor(0, 0, 0, 0))  
    return item

def spectrum_to_dict(spectrums, is_map=False):
    """Custom 'save' method to save 'Spectrum' object in a dictionary"""
    spectrums_data = spectrums.save()
    # Iterate over the saved spectrums data and update x0 and y0
    for i, spectrum in enumerate(spectrums):
        spectrum_dict = {
            "is_corrected": spectrum.is_corrected,
            "correction_value": spectrum.correction_value
        }
        
        # Save x0 and y0 only if it's not a map
        if not is_map:
            spectrum_dict.update({
                "x0": compress(spectrum.x0),
                "y0": compress(spectrum.y0)
            })
        
        # Update the spectrums_data with the new dictionary values
        spectrums_data[i].update(spectrum_dict)
    return spectrums_data

def dict_to_spectrum(spectrum, spectrum_data, is_map=True, maps=None):
    """Set attributes of Spectrum object from JSON dict"""
    spectrum.set_attributes(spectrum_data)
    
    # Set additional attributes
    spectrum.is_corrected = spectrum_data.get('is_corrected', False)
    spectrum.correction_value = spectrum_data.get('correction_value', 0)
    
    if is_map: 
        if maps is None:
            raise ValueError("maps must be provided when map=True.")
        
        # Retrieve map_name and coord from spectrum.fname
        fname = spectrum.fname
        map_name, coord_str = fname.rsplit('_', 1)
        coord_str = coord_str.strip('()')  # Remove parentheses
        coord = tuple(map(float, coord_str.split(',')))  # Convert to float tuple
        
        # Retrieve x0 and y0 from the corresponding map_df using map_name and coord
        if map_name in maps:
            map_df = maps[map_name]
            map_df = map_df.iloc[:, :-1]  # Drop the last column from map_df (NaN)
            coord_x, coord_y = coord

            row = map_df[(map_df['X'] == coord_x) & (map_df['Y'] == coord_y)]
            
            if not row.empty:
                spectrum.x0 = map_df.columns[2:].astype(float).values  
                spectrum.y0 = row.iloc[0, 2:].values  
            else:
                spectrum.x0 = None
                spectrum.y0 = None
        else:
            spectrum.x0 = None
            spectrum.y0 = None
    else:
        # Handle single spectrum case
        if 'x0' in spectrum_data:
            spectrum.x0 = decompress(spectrum_data['x0'], dtype=np.float64)
        if 'y0' in spectrum_data:
            spectrum.y0 = decompress(spectrum_data['y0'], dtype=np.float64)

def baseline_to_dict(spectrum):
    dict_baseline = dict(vars(spectrum.baseline).items())
    return dict_baseline

def dict_to_baseline(dict_baseline, spectrums):
    for spectrum in spectrums:
        # Create a fresh BaselineModel instance
        new_baseline =  BaseLine()
        for key, value in dict_baseline.items():
            setattr(new_baseline, key, deepcopy(value))
        spectrum.baseline = new_baseline

def rgba_to_named_color(rgba):
    """Convert RGBA tuple to a named color string."""
    # Check if the exact RGBA tuple exists in the dictionary
    rgba_tuple = tuple(rgba)
    if rgba_tuple in rgba_to_named_color_dict:
        return rgba_to_named_color_dict[rgba_tuple]
    else:
        # If exact match is not found, return the closest color name
        return mcolors.to_hex(rgba)  # Use hex as fallback if needed


def show_alert(message):
    """Show alert"""
    msg_box = QMessageBox()
    msg_box.setIcon(QMessageBox.Warning)
    msg_box.setWindowTitle("Alert")
    msg_box.setText(message)
    msg_box.exec_()


def clear_layout(layout):
    """To clear a given layout"""
    if layout is not None:
        while layout.count():
            item = layout.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()

def view_df(tabWidget, df, simplified_df=False, fill_colors=True):
    """View selected dataframe"""
    df_viewer = QDialog(tabWidget.parent())
    df_viewer.setWindowTitle("DataFrame Viewer")
    df_viewer.setWindowFlags(df_viewer.windowFlags() | Qt.WindowCloseButtonHint)
    layout = QVBoxLayout(df_viewer)
    layout.setContentsMargins(0, 0, 0, 0)
    dataframe_table = DataframeTableWidget(layout)
    
    if simplified_df:
        # Show a simplified version with first/last 50 rows and first/last 30 columns
        row_subset = pd.concat([df.head(50), df.tail(50)])  # First 50 and last 50 rows
        col_subset = pd.concat([row_subset.iloc[:, :30], row_subset.iloc[:, -30:]], axis=1)  # First 30 and last 30 columns
        if fill_colors:
            dataframe_table.show(col_subset)
        else:
            dataframe_table.show(col_subset, fill_colors=False)
    else:
        # Show the full dataframe
        if fill_colors:
            dataframe_table.show(df)
        else:
            dataframe_table.show(df, fill_colors=False)
    df_viewer.setLayout(layout)
    df_viewer.exec_()

def save_df_to_excel(save_path, df):
    """Saves a DataFrame to an Excel file with colored columns based on prefixes."""
    if not save_path:
        return False, "No save path provided."

    try:
        if df.empty:
            return False, "DataFrame is empty. Nothing to save."
        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Results')
            workbook = writer.book
            worksheet = writer.sheets['Results']
            palette = ['#bda16d', '#a27ba0', '#cb5b12', '#23993b', '#008281', '#147ce4']
            prefix_colors = {}

            # Apply colors based on column prefixes
            for col_idx, col_name in enumerate(df.columns, start=1):
                prefix = col_name.split("_")[0] if "_" in col_name else col_name
                if prefix not in prefix_colors:
                    color_index = len(prefix_colors) % len(palette)
                    prefix_colors[prefix] = PatternFill(start_color=palette[color_index][1:], 
                                                        end_color=palette[color_index][1:], 
                                                        fill_type='solid')
                for row in range(2, len(df) + 2):  # Apply the fill color to the entire column
                    worksheet.cell(row=row, column=col_idx).fill = prefix_colors[prefix]

        return True, "DataFrame saved successfully."
    
    except Exception as e:
        return False, f"Error when saving DataFrame: {str(e)}"

def copy_fig_to_clb(canvas, size_ratio=None):
    """Copy matplotlib figure canvas to clipboard with optional resizing."""
    current_os = platform.system()

    if canvas:
        figure = canvas.figure
        # Resize the figure if a size_ratio is provided
        if size_ratio:
            # Save the original size to restore later
            original_size = figure.get_size_inches()
            figure.set_size_inches(size_ratio, forward=True)  
            canvas.draw()  
            figure.tight_layout()

        if current_os == 'Darwin':  # macOS
            buf = BytesIO()
            canvas.print_png(buf)
            buf.seek(0)
            image = Image.open(buf)
            img_size = image.size
            png_data = buf.getvalue()
            image_rep = AppKit.NSBitmapImageRep.alloc().initWithData_(
                AppKit.NSData.dataWithBytes_length_(png_data, len(png_data))
            )
            ns_image = AppKit.NSImage.alloc().initWithSize_((img_size[0], img_size[1]))
            ns_image.addRepresentation_(image_rep)
            pasteboard = AppKit.NSPasteboard.generalPasteboard()
            pasteboard.clearContents()
            pasteboard.writeObjects_([ns_image])

        elif current_os == 'Windows':
            with BytesIO() as buf:
                figure.savefig(buf, format='png', dpi=300) 
                data = buf.getvalue()
            format_id = win32clipboard.RegisterClipboardFormat('PNG')
            win32clipboard.OpenClipboard()
            win32clipboard.EmptyClipboard()
            win32clipboard.SetClipboardData(format_id, data)
            win32clipboard.CloseClipboard()

        else:
            QMessageBox.critical(None, "Error", f"Unsupported OS: {current_os}")

        # Restore the original figure size if resized
        if size_ratio:
            figure.set_size_inches(original_size, forward=True)
            canvas.draw()  
    else:
        QMessageBox.critical(None, "Error", "No plot to copy.")
        
        
def calc_area(model_name, params):
        # params: dict with parameter values like ampli, fwhm, alpha, etc.
        if model_name == "Gaussian":
            ampli = params.get('ampli')
            fwhm = params.get('fwhm')
            if ampli is not None and fwhm is not None:
                return ampli * fwhm * np.sqrt(np.pi / (4 * np.log(2)))
        elif model_name == "Lorentzian":
            ampli = params.get('ampli')
            fwhm = params.get('fwhm')
            if ampli is not None and fwhm is not None:
                return np.pi * ampli * fwhm / 2
        elif model_name == "PseudoVoigt":
            ampli = params.get('ampli')
            fwhm = params.get('fwhm')
            alpha = params.get('alpha', 0.5)
            if ampli is not None and fwhm is not None:
                gauss_area = ampli * fwhm * np.sqrt(np.pi / (4 * np.log(2)))
                lorentz_area = np.pi * ampli * fwhm / 2
                return alpha * gauss_area + (1 - alpha) * lorentz_area
        elif model_name == "GaussianAsym":
            # asymmetric Gaussian has left and right FWHM
            ampli = params.get('ampli')
            fwhm_l = params.get('fwhm_l')
            fwhm_r = params.get('fwhm_r')
            if ampli is not None and fwhm_l is not None and fwhm_r is not None:
                area_l = (ampli * fwhm_l * np.sqrt(np.pi / (4 * np.log(2))))/2
                area_r = (ampli * fwhm_r * np.sqrt(np.pi / (4 * np.log(2))))/2
                return area_l + area_r
        elif model_name == "LorentzianAsym":
            ampli = params.get('ampli')
            fwhm_l = params.get('fwhm_l')
            fwhm_r = params.get('fwhm_r')
            if ampli is not None and fwhm_l is not None and fwhm_r is not None:
                area_l = (np.pi * ampli * fwhm_l / 2)/2
                area_r = (np.pi * ampli * fwhm_r / 2)/2
                return area_l + area_r
        return None  # default if parameters missing
    
    