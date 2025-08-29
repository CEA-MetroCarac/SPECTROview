# module to contain all common utilites / methodes used across the app

import base64
import zlib
import numpy as np
import pandas as pd
import matplotlib.colors as mcolors
import platform
from io import BytesIO
from PIL import Image

from copy import deepcopy
from openpyxl.styles import PatternFill

from PySide6.QtWidgets import QListWidgetItem, QMessageBox, QDialog, QVBoxLayout
from PySide6.QtGui import Qt, QColor

from fitspy.core.baseline import BaseLine

from spectroview.components.widget_dataframetable import DataframeTableWidget

if platform.system() == 'Darwin':
    import AppKit 
if platform.system() == 'Windows':
    import win32clipboard

def compress(array):
    """Compress and encode a numpy array to a base64 string."""
    compressed = zlib.compress(array.tobytes())
    encoded = base64.b64encode(compressed).decode('utf-8')
    return encoded


def decompress(data, dtype):
    """Decode and decompress a base64 string to a numpy array."""
    decoded = base64.b64decode(data.encode('utf-8'))
    decompressed = zlib.decompress(decoded)
    return np.frombuffer(decompressed, dtype=dtype)

def plot_baseline_dynamically(ax, spectrum):
    """Evaluate and plot baseline points and line dynamically"""
    if not spectrum.baseline.is_subtracted:
        x_bl = spectrum.x
        y_bl = spectrum.y if spectrum.baseline.attached else None
        if len(spectrum.baseline.points[0]) == 0:
            return
        # Clear any existing baseline plot
        for line in ax.lines:
            if line.get_label() == "Baseline":
                line.remove()
        # Evaluate the baseline
        attached = spectrum.baseline.attached
        baseline_values = spectrum.baseline.eval(x_bl, y_bl,
                                                    attached=attached)
        ax.plot(x_bl, baseline_values, 'r')

        # Plot the attached baseline points
        if spectrum.baseline.attached and y_bl is not None:
            attached_points = spectrum.baseline.attached_points(x_bl, y_bl)
            ax.plot(attached_points[0], attached_points[1], 'ko',
                    mfc='none')
        else:
            ax.plot(spectrum.baseline.points[0],
                    spectrum.baseline.points[1], 'ko', mfc='none', ms=5)

def populate_spectrum_listbox(spectrum, spectrum_name, checked_states):
    """ Populate the listbox with spectrums with colors"""
    item = QListWidgetItem(spectrum_name)            
    item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
    item.setCheckState(checked_states.get(spectrum_name, Qt.Checked))

    if spectrum.baseline.is_subtracted:
        if not hasattr(spectrum.result_fit, 'success'):
            item.setBackground(QColor("red"))
        elif spectrum.result_fit.success:
            item.setBackground(QColor("green"))
        else:
            item.setBackground(QColor("orange"))
    else:
        item.setBackground(QColor(0, 0, 0, 0))  
    return item

def spectrum_to_dict(spectrums, is_map=False):
    """Custom 'save' method to save 'Spectrum' object in a dictionary"""
    spectrums_data = spectrums.save()
    # Iterate over the saved spectrums data and update x0 and y0
    for i, spectrum in enumerate(spectrums):
        spectrum_dict = {
            "is_corrected": spectrum.is_corrected,
            "correction_value": spectrum.correction_value
        }
        
        # Save x0 and y0 only if it's not a map
        if not is_map:
            spectrum_dict.update({
                "x0": compress(spectrum.x0),
                "y0": compress(spectrum.y0)
            })
        
        # Update the spectrums_data with the new dictionary values
        spectrums_data[i].update(spectrum_dict)
    return spectrums_data

def dict_to_spectrum(spectrum, spectrum_data, is_map=True, maps=None):
    """Set attributes of Spectrum object from JSON dict"""
    spectrum.set_attributes(spectrum_data)
    
    # Set additional attributes
    spectrum.is_corrected = spectrum_data.get('is_corrected', False)
    spectrum.correction_value = spectrum_data.get('correction_value', 0)
    
    if is_map: 
        if maps is None:
            raise ValueError("maps must be provided when map=True.")
        
        # Retrieve map_name and coord from spectrum.fname
        fname = spectrum.fname
        map_name, coord_str = fname.rsplit('_', 1)
        coord_str = coord_str.strip('()')  # Remove parentheses
        coord = tuple(map(float, coord_str.split(',')))  # Convert to float tuple
        
        # Retrieve x0 and y0 from the corresponding map_df using map_name and coord
        if map_name in maps:
            map_df = maps[map_name]
            map_df = map_df.iloc[:, :-1]  # Drop the last column from map_df (NaN)
            coord_x, coord_y = coord

            row = map_df[(map_df['X'] == coord_x) & (map_df['Y'] == coord_y)]
            
            if not row.empty:
                spectrum.x0 = map_df.columns[2:].astype(float).values  
                spectrum.y0 = row.iloc[0, 2:].values  
            else:
                spectrum.x0 = None
                spectrum.y0 = None
        else:
            spectrum.x0 = None
            spectrum.y0 = None
    else:
        # Handle single spectrum case
        if 'x0' in spectrum_data:
            spectrum.x0 = decompress(spectrum_data['x0'], dtype=np.float64)
        if 'y0' in spectrum_data:
            spectrum.y0 = decompress(spectrum_data['y0'], dtype=np.float64)

def baseline_to_dict(spectrum):
    dict_baseline = dict(vars(spectrum.baseline).items())
    return dict_baseline

def dict_to_baseline(dict_baseline, spectrums):
    for spectrum in spectrums:
        # Create a fresh BaselineModel instance
        new_baseline =  BaseLine()
        for key, value in dict_baseline.items():
            setattr(new_baseline, key, deepcopy(value))
        spectrum.baseline = new_baseline


# Define a dictionary mapping RGBA tuples to named colors
rgba_to_named_color_dict = {mcolors.to_rgba(color_name): color_name for
                            color_name in mcolors.CSS4_COLORS}

def rgba_to_named_color(rgba):
    """Convert RGBA tuple to a named color string."""
    # Check if the exact RGBA tuple exists in the dictionary
    rgba_tuple = tuple(rgba)
    if rgba_tuple in rgba_to_named_color_dict:
        return rgba_to_named_color_dict[rgba_tuple]
    else:
        # If exact match is not found, return the closest color name
        return mcolors.to_hex(rgba)  # Use hex as fallback if needed


def show_alert(message):
    """Show alert"""
    msg_box = QMessageBox()
    msg_box.setIcon(QMessageBox.Warning)
    msg_box.setWindowTitle("Alert")
    msg_box.setText(message)
    msg_box.exec_()


def clear_layout(layout):
    """To clear a given layout"""
    if layout is not None:
        while layout.count():
            item = layout.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()

def view_df(tabWidget, df, simplified_df=False, fill_colors=True):
    """View selected dataframe"""
    df_viewer = QDialog(tabWidget.parent())
    df_viewer.setWindowTitle("DataFrame Viewer")
    df_viewer.setWindowFlags(df_viewer.windowFlags() | Qt.WindowCloseButtonHint)
    layout = QVBoxLayout(df_viewer)
    layout.setContentsMargins(0, 0, 0, 0)
    dataframe_table = DataframeTableWidget(layout)
    
    if simplified_df:
        # Show a simplified version with first/last 50 rows and first/last 30 columns
        row_subset = pd.concat([df.head(50), df.tail(50)])  # First 50 and last 50 rows
        col_subset = pd.concat([row_subset.iloc[:, :30], row_subset.iloc[:, -30:]], axis=1)  # First 30 and last 30 columns
        if fill_colors:
            dataframe_table.show(col_subset)
        else:
            dataframe_table.show(col_subset, fill_colors=False)
    else:
        # Show the full dataframe
        if fill_colors:
            dataframe_table.show(df)
        else:
            dataframe_table.show(df, fill_colors=False)
    df_viewer.setLayout(layout)
    df_viewer.exec_()

def save_df_to_excel(save_path, df):
    """Saves a DataFrame to an Excel file with colored columns based on prefixes."""
    if not save_path:
        return False, "No save path provided."

    try:
        if df.empty:
            return False, "DataFrame is empty. Nothing to save."
        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Results')
            workbook = writer.book
            worksheet = writer.sheets['Results']
            palette = ['#bda16d', '#a27ba0', '#cb5b12', '#23993b', '#008281', '#147ce4']
            prefix_colors = {}

            # Apply colors based on column prefixes
            for col_idx, col_name in enumerate(df.columns, start=1):
                prefix = col_name.split("_")[0] if "_" in col_name else col_name
                if prefix not in prefix_colors:
                    color_index = len(prefix_colors) % len(palette)
                    prefix_colors[prefix] = PatternFill(start_color=palette[color_index][1:], 
                                                        end_color=palette[color_index][1:], 
                                                        fill_type='solid')
                for row in range(2, len(df) + 2):  # Apply the fill color to the entire column
                    worksheet.cell(row=row, column=col_idx).fill = prefix_colors[prefix]

        return True, "DataFrame saved successfully."
    
    except Exception as e:
        return False, f"Error when saving DataFrame: {str(e)}"

def copy_fig_to_clb(canvas, size_ratio=None):
    """Copy matplotlib figure canvas to clipboard with optional resizing."""
    current_os = platform.system()

    if canvas:
        figure = canvas.figure
        # Resize the figure if a size_ratio is provided
        if size_ratio:
            # Save the original size to restore later
            original_size = figure.get_size_inches()
            figure.set_size_inches(size_ratio, forward=True)  
            canvas.draw()  
            figure.tight_layout()

        if current_os == 'Darwin':  # macOS
            buf = BytesIO()
            canvas.print_png(buf)
            buf.seek(0)
            image = Image.open(buf)
            img_size = image.size
            png_data = buf.getvalue()
            image_rep = AppKit.NSBitmapImageRep.alloc().initWithData_(
                AppKit.NSData.dataWithBytes_length_(png_data, len(png_data))
            )
            ns_image = AppKit.NSImage.alloc().initWithSize_((img_size[0], img_size[1]))
            ns_image.addRepresentation_(image_rep)
            pasteboard = AppKit.NSPasteboard.generalPasteboard()
            pasteboard.clearContents()
            pasteboard.writeObjects_([ns_image])

        elif current_os == 'Windows':
            with BytesIO() as buf:
                figure.savefig(buf, format='png', dpi=300) 
                data = buf.getvalue()
            format_id = win32clipboard.RegisterClipboardFormat('PNG')
            win32clipboard.OpenClipboard()
            win32clipboard.EmptyClipboard()
            win32clipboard.SetClipboardData(format_id, data)
            win32clipboard.CloseClipboard()

        else:
            QMessageBox.critical(None, "Error", f"Unsupported OS: {current_os}")

        # Restore the original figure size if resized
        if size_ratio:
            figure.set_size_inches(original_size, forward=True)
            canvas.draw()  
    else:
        QMessageBox.critical(None, "Error", "No plot to copy.")
        
        
def calc_area(model_name, params):
        # params: dict with parameter values like ampli, fwhm, alpha, etc.
        if model_name == "Gaussian":
            ampli = params.get('ampli')
            fwhm = params.get('fwhm')
            if ampli is not None and fwhm is not None:
                return ampli * fwhm * np.sqrt(np.pi / (4 * np.log(2)))
        elif model_name == "Lorentzian":
            ampli = params.get('ampli')
            fwhm = params.get('fwhm')
            if ampli is not None and fwhm is not None:
                return np.pi * ampli * fwhm / 2
        elif model_name == "PseudoVoigt":
            ampli = params.get('ampli')
            fwhm = params.get('fwhm')
            alpha = params.get('alpha', 0.5)
            if ampli is not None and fwhm is not None:
                gauss_area = ampli * fwhm * np.sqrt(np.pi / (4 * np.log(2)))
                lorentz_area = np.pi * ampli * fwhm / 2
                return alpha * gauss_area + (1 - alpha) * lorentz_area
        elif model_name == "GaussianAsym":
            # asymmetric Gaussian has left and right FWHM
            ampli = params.get('ampli')
            fwhm_l = params.get('fwhm_l')
            fwhm_r = params.get('fwhm_r')
            if ampli is not None and fwhm_l is not None and fwhm_r is not None:
                area_l = (ampli * fwhm_l * np.sqrt(np.pi / (4 * np.log(2))))/2
                area_r = (ampli * fwhm_r * np.sqrt(np.pi / (4 * np.log(2))))/2
                return area_l + area_r
        elif model_name == "LorentzianAsym":
            ampli = params.get('ampli')
            fwhm_l = params.get('fwhm_l')
            fwhm_r = params.get('fwhm_r')
            if ampli is not None and fwhm_l is not None and fwhm_r is not None:
                area_l = (np.pi * ampli * fwhm_l / 2)/2
                area_r = (np.pi * ampli * fwhm_r / 2)/2
                return area_l + area_r
        return None  # default if parameters missing