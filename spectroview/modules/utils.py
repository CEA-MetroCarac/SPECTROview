# module to contain all common utilites / methodes used across the app
import os
import base64
import markdown
import zlib
import numpy as np
import pandas as pd
import platform

import matplotlib.colors as mcolors
import matplotlib.cm as cm
from matplotlib.figure import Figure
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT

from io import BytesIO
from PIL import Image
from threading import Thread
from multiprocessing import Queue
from copy import deepcopy
from openpyxl.styles import PatternFill

from fitspy.core.spectrum import Spectrum as FitspySpectrum
from fitspy.core.spectra import Spectra as FitspySpectra
from fitspy.core.baseline import BaseLine
from fitspy.core.utils_mp import fit_mp

from spectroview import PALETTE, DEFAULT_COLORS
from spectroview.modules.df_table import DataframeTable

from PySide6.QtWidgets import QDialog, QTableWidgetItem, QVBoxLayout,  QTextBrowser, \
    QComboBox, QListWidgetItem, QMessageBox, QDialog, QVBoxLayout, QListWidget, QAbstractItemView
    
from PySide6.QtCore import Signal, QThread, Qt, QSize
from PySide6.QtGui import QPalette, QColor, QTextCursor, QIcon, Qt, QPixmap, QImage

if platform.system() == 'Darwin':
    import AppKit 
if platform.system() == 'Windows':
    import win32clipboard
    
    
class CustomizedPalette(QComboBox):
    """Custom QComboBox to show color palette previews along with their names."""
    def __init__(self, palette_list=None, parent=None, icon_size=(99, 12)):
        super().__init__(parent)
        self.icon_width, self.icon_height = icon_size
        self.setIconSize(QSize(*icon_size))
        self.setMinimumWidth(100)

        self.palette_list = palette_list or PALETTE
        self._populate_with_previews()

    def _populate_with_previews(self):
        self.clear()
        for cmap_name in self.palette_list:
            icon = QIcon(self._create_colormap_preview(cmap_name))
            self.addItem(icon, cmap_name)

    def _create_colormap_preview(self, cmap_name):
        """Generate a horizontal gradient preview image for the colormap."""
        width, height = self.icon_width, self.icon_height
        gradient = np.linspace(0, 1, 20).reshape(1, -1)

        fig = Figure(figsize=(width / 100, height / 100), dpi=100)
        canvas = FigureCanvas(fig)
        ax = fig.add_axes([0, 0, 1, 1], frameon=False)
        ax.imshow(gradient, aspect='auto', cmap=cm.get_cmap(cmap_name))
        ax.set_axis_off()
        canvas.draw()

        image = np.array(canvas.buffer_rgba())
        qimage = QImage(image.data, image.shape[1], image.shape[0],
                        QImage.Format_RGBA8888)
        return QPixmap.fromImage(qimage)

    def get_selected_palette(self):
        return self.currentText()

class Spectrum(FitspySpectrum):
    """Customized of Spectrum class."""
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.label = None   # user-defined legend label
        self.color = None   # user-defined color
        
        self.xcorrection_value = 0   # peak position correction using reference value
                    
    def reinit(self):
        """ Reinitialize the main attributes """
        self.range_min = None
        self.range_max = None
        self.x = self.x0.copy()
        self.y = self.y0.copy()
        self.weights = self.weights0.copy() if self.weights0 is not None else None
        self.outliers_limit = None
        self.normalize = False
        self.normalize_range_min = None
        self.normalize_range_max = None
        self.remove_models()
        self.result_fit = lambda: None
        self.baseline.reinit()
        self.baseline.mode = "Linear"
        
        self.color = None
        self.label = None

    def preprocess(self):
        """ Preprocess the spectrum """
        self.load_profile(self.fname)
        self.apply_range()
        self.eval_baseline()
        self.subtract_baseline()
        self.normalization()

    def apply_xcorrection(self, new_xcorr_value=None):
        """ Apply peak position correction """
        # Undo existing correction if needed
        if self.xcorrection_value != 0:
            self.undo_xcorrection()

        # If user provides a new correction, update the value
        if new_xcorr_value is not None:
            self.xcorrection_value = new_xcorr_value

        # Apply correction
        if self.xcorrection_value != 0:
            self.x0 = self.x0 + self.xcorrection_value
            self.x = self.x + self.xcorrection_value

    def undo_xcorrection(self):
        """Undo peak position correction (restore original x and x0)."""
        if self.xcorrection_value != 0:
            self.x0 = self.x0 - self.xcorrection_value
            self.x = self.x - self.xcorrection_value
            self.xcorrection_value = 0
        
class Spectra(FitspySpectra):
    """Customized Spectra class"""
    
    def apply_model(self, model_dict, fnames=None, ncpus=1,show_progressbar=True):
        """ Apply 'model' to all or part of the spectra."""
        if fnames is None:
            fnames = self.fnames

        spectra = []
        for fname in fnames:
            spectrum, _ = self.get_objects(fname)
            
            # Customize the model_dict for this spectrum
            custom_model = deepcopy(model_dict)
            if hasattr(spectrum, "xcorrection_value"):  # reassign current xcorrection_value
                custom_model["xcorrection_value"] = spectrum.xcorrection_value
            if hasattr(spectrum, "label"):  
                custom_model["label"] = spectrum.label
            if hasattr(spectrum, "color"):  
                custom_model["color"] = spectrum.color

            spectrum.set_attributes(custom_model)
            spectrum.fname = fname  # reassign the correct fname
            spectra.append(spectrum)

        self.pbar_index = 0

        queue_incr = Queue()
        args = (queue_incr, len(fnames), ncpus, show_progressbar)
        thread = Thread(target=self.progressbar, args=args)
        thread.start()

        if ncpus == 1:
            for spectrum in spectra:
                spectrum.preprocess()
                spectrum.fit()
                queue_incr.put(1)
        else:
            fit_mp(spectra, ncpus, queue_incr)
        thread.join()     
        
class FitThread(QThread):
    """ Class to perform fitting in a separate Thread """
    progress_changed = Signal(int)
    def __init__(self, spectrums, fit_model, fnames, ncpus=1):
        super().__init__()
        self.spectrums = spectrums
        self.fit_model = fit_model
        self.fnames = fnames
        self.ncpus = ncpus

    def run(self):
        fit_model = deepcopy(self.fit_model)
        self.spectrums.apply_model(fit_model, fnames=self.fnames,
                                   ncpus=self.ncpus, show_progressbar=False)

        self.progress_changed.emit(100)

class CustomizedListWidget(QListWidget):
    """
    Customized QListWidget with drag-and-drop functionality for rearranging
    items.
    """
    items_reordered = Signal()
    files_dropped = Signal(list) 

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)  # Enable external drag-drop
        self.setDragDropMode(QListWidget.InternalMove)
        self.setSelectionMode(QAbstractItemView.ExtendedSelection)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            super().dragEnterEvent(event)
            
    def dragMoveEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            super().dragMoveEvent(event)
            
    def dropEvent(self, event):
        """Overrides the dropEvent method to emit the items_reordered signal"""
        if event.mimeData().hasUrls():
            file_paths = [url.toLocalFile() for url in event.mimeData().urls()]
            self.files_dropped.emit(file_paths)  # emit signal with file list
            event.acceptProposedAction()
        else:
            super().dropEvent(event)
            self.items_reordered.emit()

def view_markdown(ui, title, fname, x, y, working_folder):
    """To convert MD file to html format and display them in GUI"""
    with open(fname, 'r', encoding='utf-8') as f:
        markdown_content = f.read()
    html_content = markdown.markdown(markdown_content)
    DIRNAME = os.path.dirname(__file__)
    html_content = html_content.replace('src="',
                                        f'src="'
                                        f'{os.path.join(DIRNAME, working_folder)}')
    about_dialog = QDialog(ui)
    about_dialog.setWindowTitle(title)
    about_dialog.resize(x, y)
    text_browser = QTextBrowser(about_dialog)
    text_browser.setAlignment(Qt.AlignLeft | Qt.AlignTop)
    text_browser.setOpenExternalLinks(True)
    text_browser.setHtml(html_content)
    layout = QVBoxLayout(about_dialog)
    layout.addWidget(text_browser)
    about_dialog.setLayout(layout)
    about_dialog.show()
                    
def dark_palette():
        """Palette color for dark mode of the appli's GUI"""
        dark_palette = QPalette()
        dark_palette.setColor(QPalette.Window, QColor(70, 70, 70))
        dark_palette.setColor(QPalette.WindowText, Qt.white)
        dark_palette.setColor(QPalette.AlternateBase, QColor(45, 45, 45))
        dark_palette.setColor(QPalette.ToolTipBase, QColor(255, 255, 220))
        dark_palette.setColor(QPalette.ToolTipText, Qt.black)
        dark_palette.setColor(QPalette.Text, Qt.white)
        dark_palette.setColor(QPalette.Button, QColor(64, 64, 64))
        dark_palette.setColor(QPalette.ButtonText, Qt.white)
        dark_palette.setColor(QPalette.BrightText, Qt.red)
        dark_palette.setColor(QPalette.Link, QColor(42, 130, 218))
        dark_palette.setColor(QPalette.Highlight, QColor(42, 130, 218))
        dark_palette.setColor(QPalette.HighlightedText, Qt.white)
        dark_palette.setColor(QPalette.PlaceholderText, QColor(140, 140, 140))
        dark_palette.setColor(QPalette.Base, QColor(60, 60, 60))  # Background color for QMenu
        
        return dark_palette

def light_palette():
    """Palette color for light mode of the appli's GUI"""
    light_palette = QPalette()
    light_palette.setColor(QPalette.Window, QColor(225, 225, 225))
    light_palette.setColor(QPalette.WindowText, Qt.black)
    light_palette.setColor(QPalette.AlternateBase, QColor(230, 230, 230))
    light_palette.setColor(QPalette.ToolTipBase, QColor(255, 255, 220))
    light_palette.setColor(QPalette.ToolTipText, Qt.black)
    light_palette.setColor(QPalette.Text, Qt.black)
    light_palette.setColor(QPalette.Button, QColor(230, 230, 230))
    light_palette.setColor(QPalette.ButtonText, Qt.black)
    light_palette.setColor(QPalette.BrightText, Qt.red)
    light_palette.setColor(QPalette.Link, QColor(42, 130, 218))
    light_palette.setColor(QPalette.Highlight, QColor(42, 130, 218))
    light_palette.setColor(QPalette.HighlightedText, Qt.black)
    light_palette.setColor(QPalette.PlaceholderText, QColor(150, 150, 150))
    light_palette.setColor(QPalette.Base, QColor(240, 240, 240))  # Menu background color

    return light_palette
                
def view_text(ui, title, text):
        """ Create a QTextBrowser to display a text content"""
        report_viewer = QDialog(ui)
        report_viewer.setWindowTitle(title)
        report_viewer.setGeometry(100, 100, 800, 600)
        text_browser = QTextBrowser(report_viewer)
        text_browser.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        text_browser.setOpenExternalLinks(True)
        text_browser.setPlainText(text)
        text_browser.moveCursor(QTextCursor.Start)
        layout = QVBoxLayout(report_viewer)
        layout.addWidget(text_browser)
        report_viewer.show()            
            
def quadrant(row):
        """Define 4 quadrant of a wafer"""
        if row['X'] < 0 and row['Y'] < 0:
            return 'Q1'
        elif row['X'] < 0 and row['Y'] > 0:
            return 'Q2'
        elif row['X'] > 0 and row['Y'] > 0:
            return 'Q3'
        elif row['X'] > 0 and row['Y'] < 0:
            return 'Q4'
        else:
            return np.nan

def zone(row, radius):
    """Define 3 zones (Center, Mid-Radius, Edge)"""
    r = radius
    x = row['X']
    y = row['Y']
    distance_to_center = np.sqrt(x ** 2 + y ** 2)
    if distance_to_center <= r * 0.35:
        return 'Center'
    elif distance_to_center > r * 0.35 and distance_to_center < r * 0.8:
        return 'Mid-Radius'
    elif distance_to_center >= 0.8 * r:
        return 'Edge'
    else:
        return np.nan
    
def replace_peak_labels(fit_model, param):
        """Replace prefix 'm01' of peak model by labels designed by user"""
        peak_labels = fit_model["peak_labels"]
        if "_" in param:
            prefix, param = param.split("_", 1)  
            # Convert prefix to peak_label
            peak_index = int(prefix[1:]) - 1
            if 0 <= peak_index < len(peak_labels):
                peak_label = peak_labels[peak_index]
                return f"{param}_{peak_label}"
        return param
    
def compress(array):
    """Compress and encode a numpy array to a base64 string."""
    compressed = zlib.compress(array.tobytes())
    encoded = base64.b64encode(compressed).decode('utf-8')
    return encoded


def decompress(data, dtype):
    """Decode and decompress a base64 string to a numpy array."""
    decoded = base64.b64decode(data.encode('utf-8'))
    decompressed = zlib.decompress(decoded)
    return np.frombuffer(decompressed, dtype=dtype)

def plot_baseline_dynamically(ax, spectrum):
    """Evaluate and plot baseline points and line dynamically"""
    if not spectrum.baseline.is_subtracted:
        x_bl = spectrum.x
        y_bl = spectrum.y if spectrum.baseline.attached else None
        if len(spectrum.baseline.points[0]) == 0:
            return
        # Clear any existing baseline plot
        for line in ax.lines:
            if line.get_label() == "Baseline":
                line.remove()
        # Evaluate the baseline
        attached = spectrum.baseline.attached
        baseline_values = spectrum.baseline.eval(x_bl, y_bl,
                                                    attached=attached)
        ax.plot(x_bl, baseline_values, 'r')

        # Plot the attached baseline points
        if spectrum.baseline.attached and y_bl is not None:
            attached_points = spectrum.baseline.attached_points(x_bl, y_bl)
            ax.plot(attached_points[0], attached_points[1], 'ko',
                    mfc='none')
        else:
            ax.plot(spectrum.baseline.points[0],
                    spectrum.baseline.points[1], 'ko', mfc='none', ms=5)

def populate_spectrum_listbox(spectrum, spectrum_name, checked_states):
    """ Populate the listbox with spectrums with colors"""
    item = QListWidgetItem(spectrum_name)            
    item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
    item.setCheckState(checked_states.get(spectrum_name, Qt.Checked))

    if spectrum.baseline.is_subtracted:
        if not hasattr(spectrum.result_fit, 'success'):
            item.setBackground(QColor("red"))
        elif spectrum.result_fit.success:
            item.setBackground(QColor("green"))
        else:
            item.setBackground(QColor("orange"))
    else:
        item.setBackground(QColor(0, 0, 0, 0))  
    return item

def spectrum_to_dict(spectrums, is_map=False):
    """Custom 'save' method to save 'Spectrum' object in a dictionary"""
    spectrums_data = spectrums.save()
    # Iterate over the saved spectrums data and update x0 and y0
    for i, spectrum in enumerate(spectrums):
        spectrum_dict = {}
        
        # Save x0 and y0 only if it's not a 2DMAP
        if not is_map:
            spectrum_dict.update({
                "x0": compress(spectrum.x0),
                "y0": compress(spectrum.y0)
            })
        
        # Update the spectrums_data with the new dictionary values
        spectrums_data[i].update(spectrum_dict)
    return spectrums_data

def dict_to_spectrum(spectrum, spectrum_data, is_map=True, maps=None):
    """Set attributes of Spectrum object from JSON dict"""
    spectrum.set_attributes(spectrum_data)
    
    if is_map: 
        if maps is None:
            raise ValueError("maps must be provided when map=True.")
        
        # Retrieve map_name and coord from spectrum.fname
        fname = spectrum.fname
        map_name, coord_str = fname.rsplit('_', 1)
        coord_str = coord_str.strip('()')  # Remove parentheses
        coord = tuple(map(float, coord_str.split(',')))  # Convert to float tuple
        
        # Retrieve x0 and y0 from the corresponding map_df using map_name and coord
        if map_name in maps:
            map_df = maps[map_name]
            map_df = map_df.iloc[:, :-1]  # Drop the last column from map_df (NaN)
            coord_x, coord_y = coord

            row = map_df[(map_df['X'] == coord_x) & (map_df['Y'] == coord_y)]
            
            if not row.empty:
                x0 = map_df.columns[2:].astype(float).values  # retreive original x0
                spectrum.x0 = x0 + spectrum.xcorrection_value # apply xcorrection_value
                spectrum.y0 = row.iloc[0, 2:].values  
            else:
                spectrum.x0 = None
                spectrum.y0 = None
        else:
            spectrum.x0 = None
            spectrum.y0 = None
    else:
        # Handle single spectrum case
        if 'x0' in spectrum_data:
            spectrum.x0 = decompress(spectrum_data['x0'], dtype=np.float64)
        if 'y0' in spectrum_data:
            spectrum.y0 = decompress(spectrum_data['y0'], dtype=np.float64)

def baseline_to_dict(spectrum):
    dict_baseline = dict(vars(spectrum.baseline).items())
    return dict_baseline

def dict_to_baseline(dict_baseline, spectrums):
    for spectrum in spectrums:
        # Create a fresh BaselineModel instance
        new_baseline =  BaseLine()
        for key, value in dict_baseline.items():
            setattr(new_baseline, key, deepcopy(value))
        spectrum.baseline = new_baseline


def rgba_to_default_color(rgba, default_colors=DEFAULT_COLORS):
    """
    Convert an RGBA tuple to the closest color in DEFAULT_COLORS.
    If no DEFAULT_COLORS are given, falls back to hex.
    """
    # Convert input to RGB array
    rgb = np.array(mcolors.to_rgb(rgba)) # drops alpha

    # Compute distance to each default color
    best_color = None
    best_dist = float("inf")
    for dc in default_colors:
        dc_rgb = np.array(mcolors.to_rgb(dc))
        dist = np.linalg.norm(rgb - dc_rgb)  # Euclidean distance in RGB
        if dist < best_dist:
            best_dist = dist
            best_color = dc

    return best_color if best_color else mcolors.to_hex(rgba)

def show_alert(message):
    """Show alert"""
    msg_box = QMessageBox()
    msg_box.setIcon(QMessageBox.Warning)
    msg_box.setWindowTitle("Alert")
    msg_box.setText(message)
    msg_box.exec_()


def clear_layout(layout):
    """To clear a given layout"""
    if layout is not None:
        while layout.count():
            item = layout.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()

def view_df(tabWidget, df, simplified_df=False, fill_colors=True):
    """View selected dataframe"""
    df_viewer = QDialog(tabWidget.parent())
    df_viewer.setWindowTitle("DataFrame Viewer")
    df_viewer.setWindowFlags(df_viewer.windowFlags() | Qt.WindowCloseButtonHint)
    layout = QVBoxLayout(df_viewer)
    layout.setContentsMargins(0, 0, 0, 0)
    dataframe_table = DataframeTable(layout)
    
    if simplified_df:
        # Show a simplified version with first/last 50 rows and first/last 30 columns
        row_subset = pd.concat([df.head(50), df.tail(50)])  # First 50 and last 50 rows
        col_subset = pd.concat([row_subset.iloc[:, :30], row_subset.iloc[:, -30:]], axis=1)  # First 30 and last 30 columns
        if fill_colors:
            dataframe_table.show(col_subset)
        else:
            dataframe_table.show(col_subset, fill_colors=False)
    else:
        # Show the full dataframe
        if fill_colors:
            dataframe_table.show(df)
        else:
            dataframe_table.show(df, fill_colors=False)
    df_viewer.setLayout(layout)
    df_viewer.exec_()

def save_df_to_excel(save_path, df):
    """Saves a DataFrame to an Excel file with colored columns based on prefixes."""
    if not save_path:
        return False, "No save path provided."

    try:
        if df.empty:
            return False, "DataFrame is empty. Nothing to save."
        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Results')
            workbook = writer.book
            worksheet = writer.sheets['Results']
            palette = ['#bda16d', '#a27ba0', '#cb5b12', '#23993b', '#008281', '#147ce4']
            prefix_colors = {}

            # Apply colors based on column prefixes
            for col_idx, col_name in enumerate(df.columns, start=1):
                prefix = col_name.split("_")[0] if "_" in col_name else col_name
                if prefix not in prefix_colors:
                    color_index = len(prefix_colors) % len(palette)
                    prefix_colors[prefix] = PatternFill(start_color=palette[color_index][1:], 
                                                        end_color=palette[color_index][1:], 
                                                        fill_type='solid')
                for row in range(2, len(df) + 2):  # Apply the fill color to the entire column
                    worksheet.cell(row=row, column=col_idx).fill = prefix_colors[prefix]

        return True, "DataFrame saved successfully."
    
    except Exception as e:
        return False, f"Error when saving DataFrame: {str(e)}"

def copy_fig_to_clb(canvas, size_ratio=None):
    """Copy matplotlib figure canvas to clipboard with optional resizing."""
    current_os = platform.system()

    if canvas:
        figure = canvas.figure
        # Resize the figure if a size_ratio is provided
        if size_ratio:
            # Save the original size to restore later
            original_size = figure.get_size_inches()
            figure.set_size_inches(size_ratio, forward=True)  
            canvas.draw()  
            figure.tight_layout()

        if current_os == 'Darwin':  # macOS
            buf = BytesIO()
            canvas.print_png(buf)
            buf.seek(0)
            image = Image.open(buf)
            img_size = image.size
            png_data = buf.getvalue()
            image_rep = AppKit.NSBitmapImageRep.alloc().initWithData_(
                AppKit.NSData.dataWithBytes_length_(png_data, len(png_data))
            )
            ns_image = AppKit.NSImage.alloc().initWithSize_((img_size[0], img_size[1]))
            ns_image.addRepresentation_(image_rep)
            pasteboard = AppKit.NSPasteboard.generalPasteboard()
            pasteboard.clearContents()
            pasteboard.writeObjects_([ns_image])

        elif current_os == 'Windows':
            with BytesIO() as buf:
                figure.savefig(buf, format='png', dpi=300) 
                data = buf.getvalue()
            format_id = win32clipboard.RegisterClipboardFormat('PNG')
            win32clipboard.OpenClipboard()
            win32clipboard.EmptyClipboard()
            win32clipboard.SetClipboardData(format_id, data)
            win32clipboard.CloseClipboard()

        else:
            QMessageBox.critical(None, "Error", f"Unsupported OS: {current_os}")

        # Restore the original figure size if resized
        if size_ratio:
            figure.set_size_inches(original_size, forward=True)
            canvas.draw()  
    else:
        QMessageBox.critical(None, "Error", "No plot to copy.")
        
        
def calc_area(model_name, params):
        # params: dict with parameter values like ampli, fwhm, alpha, etc.
        if model_name == "Gaussian":
            ampli = params.get('ampli')
            fwhm = params.get('fwhm')
            if ampli is not None and fwhm is not None:
                return ampli * fwhm * np.sqrt(np.pi / (4 * np.log(2)))
        elif model_name == "Lorentzian":
            ampli = params.get('ampli')
            fwhm = params.get('fwhm')
            if ampli is not None and fwhm is not None:
                return np.pi * ampli * fwhm / 2
        elif model_name == "PseudoVoigt":
            ampli = params.get('ampli')
            fwhm = params.get('fwhm')
            alpha = params.get('alpha', 0.5)
            if ampli is not None and fwhm is not None:
                gauss_area = ampli * fwhm * np.sqrt(np.pi / (4 * np.log(2)))
                lorentz_area = np.pi * ampli * fwhm / 2
                return alpha * gauss_area + (1 - alpha) * lorentz_area
        elif model_name == "GaussianAsym":
            # asymmetric Gaussian has left and right FWHM
            ampli = params.get('ampli')
            fwhm_l = params.get('fwhm_l')
            fwhm_r = params.get('fwhm_r')
            if ampli is not None and fwhm_l is not None and fwhm_r is not None:
                area_l = (ampli * fwhm_l * np.sqrt(np.pi / (4 * np.log(2))))/2
                area_r = (ampli * fwhm_r * np.sqrt(np.pi / (4 * np.log(2))))/2
                return area_l + area_r
        elif model_name == "LorentzianAsym":
            ampli = params.get('ampli')
            fwhm_l = params.get('fwhm_l')
            fwhm_r = params.get('fwhm_r')
            if ampli is not None and fwhm_l is not None and fwhm_r is not None:
                area_l = (np.pi * ampli * fwhm_l / 2)/2
                area_r = (np.pi * ampli * fwhm_r / 2)/2
                return area_l + area_r
        return None  # default if parameters missing