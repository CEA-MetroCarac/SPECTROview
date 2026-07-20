# spectroview/viewmodel/utils.py
from PySide6.QtGui import QColor, QPixmap, QImage, QTextCursor, QIcon, QPainter

def get_tinted_icon(path: str, color_str: str) -> QIcon:
    """Return a QIcon created from the image at `path` tinted to `color_str`."""
    pixmap = QPixmap(path)
    painter = QPainter(pixmap)
    painter.setCompositionMode(QPainter.CompositionMode_SourceIn)
    painter.fillRect(pixmap.rect(), QColor(color_str))
    painter.end()
    return QIcon(pixmap)

from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QMessageBox, QApplication, QListWidgetItem,
    QDialog, QTextBrowser, QVBoxLayout
)    

import os
import datetime
import struct
import re
import numpy as np
import matplotlib
from io import BytesIO

from copy import deepcopy

from spectroview import DEFAULT_COLORS



def parse_wdf_metadata(reader):
    """Extract comprehensive metadata from WDF file's WXIS and WXCS blocks.
    
    This function parses unparsed WDF file blocks to extract instrument settings
   that are not exposed by the renishawWiRE package but are stored in the file.
    
    Args:
        reader: WDFReader object with file already opened
        
    Returns:
        dict with keys: objective_name, grating_name, exposure_time, 
                       slit_opening, slit_centre, laser_power (if available)
    """
    metadata = {}
    
    # Parse WXIS block for objective, grating, slit, and laser power
    if 'WXIS' in reader.block_info:
        uid, pos, size = reader.block_info['WXIS']
        reader.file_obj.seek(pos + 16)
        wxis_data = reader.file_obj.read(size - 16)
        
        try:
            text_data = wxis_data.decode('utf-8', errors='ignore')
            
            # Extract objective (e.g., "x100", "x50", "x20")
            for objective in ['x100', 'x50', 'x20', 'x10', 'x5', 'x2']:
                if objective in text_data:
                    metadata['objective_name'] = objective
                    break
            
            # Extract grating (e.g., "1800 l/mm")
            grating_patterns = [
                '1800 l/mm', '2400 l/mm', '1200 l/mm',
                '600 l/mm', '300 l/mm', '150 l/mm'
            ]
            for pattern in grating_patterns:
                if pattern in text_data:
                    metadata['grating_name'] = pattern
                    break
            
            # Extract SLIT OPENING and centre using labels
            # More robust: find all number-µm pairs
            all_um_values = re.findall(r'(\d+)[\xc2\xb5µ]m', text_data)
            if len(all_um_values) >= 3:
                # Skip first value (bias), take second (opening) and third (centre)
                try:
                    slit_opening = float(all_um_values[1])  # Second value
                    slit_centre = float(all_um_values[2])   # Third value
                    
                    # Sanity check: slit opening typically 10-200, centre typically 100-3000
                    if 5 <= slit_opening <= 500 and 50 <= slit_centre <= 5000:
                        metadata['slit_opening'] = slit_opening
                        metadata['slit_centre'] = slit_centre
                except (ValueError, IndexError):
                    pass
            
            # Extract LASER POWER from WXIS structured records
            # Laser power is stored as a 'u' (string) record with id=0x0500
            # Located near offset 2300-2400 with a simple numeric string value
            # Record format: u(0x75) + 2-byte-id(0x0500) + \x80 + 4-byte-strlen + string
            # Search specifically in the range 2300-2500 for a simple numeric pattern
            for i in range(2300, min(len(wxis_data) - 12, 2500)):
                if (wxis_data[i:i+1] == b'u' and wxis_data[i+3:i+4] == b'\x80'):
                    try:
                        rec_id = struct.unpack('<H', wxis_data[i+1:i+3])[0]
                        str_len = struct.unpack('<i', wxis_data[i+4:i+8])[0]
                        if rec_id == 0x0500 and 0 < str_len < 20:
                            power_str = wxis_data[i+8:i+8+str_len].decode(
                                'utf-8', errors='ignore').strip('\x00')
                            # Check if it's a simple numeric string (laser power)
                            # Should be like "0.5" or "1" (not "30.8786 Degrees")
                            if power_str and not any(c.isalpha() for c in power_str):
                                try:
                                    power_val = float(power_str)
                                    # Laser power typically 0.1% to 100%
                                    if 0.01 <= power_val <= 100:
                                        metadata['laser_power'] = power_val
                                        break
                                except ValueError:
                                    pass
                    except (struct.error, IndexError):
                        pass
            
        except Exception:
            pass
    
    # Parse WXCS block for confocal values (alternative source)
    if 'WXCS' in reader.block_info and ('slit_opening' not in metadata or 'slit_centre' not in metadata):
        uid, pos, size = reader.block_info['WXCS']
        reader.file_obj.seek(pos + 16)
        wxcs_data = reader.file_obj.read(size - 16)
        
        try:
            text_data = wxcs_data.decode('utf-8', errors='ignore')
            
            # WXCS contains "20.0µm" and "1724.7µm" for confocal opening and centre
            all_um_values = re.findall(r'(\d+\.?\d*)[\xc2\xb5µ]m', text_data)
            
            # Look for two values where one is small (~20) and one is large (~1724)
            for i in range(len(all_um_values) - 1):
                try:
                    val1 = float(all_um_values[i])
                    val2 = float(all_um_values[i+1])
                    
                    # Identify which is opening (smaller) and which is centre (larger)
                    if 5 <= val1 <= 500 and 50 <= val2 <= 5000:
                        if val1 < val2 and 'slit_opening' not in metadata:
                            metadata['slit_opening'] = val1
                            metadata['slit_centre'] = val2
                            break
                        elif val2 < val1 and 'slit_opening' not in metadata:
                            metadata['slit_opening'] = val2
                            metadata['slit_centre'] = val1
                            break
                except ValueError:
                    pass
        except Exception:
            pass
    
    # Parse WXDM block for exposure time
    # Exposure time is stored as int32 milliseconds in an 'i' record with id=0x2300
    # Record format: i(0x69) + 2-byte-id(0x2300) + \x80 + 4-byte-int-value
    if 'WXDM' in reader.block_info:
        uid, pos, size = reader.block_info['WXDM']
        reader.file_obj.seek(pos + 16)
        wxdm_data = reader.file_obj.read(size - 16)
        
        try:
            for i in range(len(wxdm_data) - 8):
                if (wxdm_data[i:i+1] == b'i'
                        and wxdm_data[i+3:i+4] == b'\x80'):
                    rec_id = struct.unpack('<H', wxdm_data[i+1:i+3])[0]
                    if rec_id == 0x2300:  # Exposure time record ID
                        val = struct.unpack('<i', wxdm_data[i+4:i+8])[0]
                        if 10 <= val <= 3600000:  # 10ms to 1 hour
                            metadata['exposure_time'] = val / 1000.0  # Convert ms to seconds
                            break
        except Exception:
            pass
    
    # Parse WDF1 block to find Creation Date
    # Priority:
    # 1. Windows FILETIME at 0x90 (End Time)
    # 2. Windows FILETIME at 0x88 (Start Time)
    # 3. Check for SYSTEMTIME structure (legacy fallback)
    # 4. File modification time (last resort)
    
    found_date = None
    
    try:
        reader.file_obj.seek(0)
        header_bytes = reader.file_obj.read(1024) # Read enough to cover header

        # Helper to decode FILETIME (100-ns intervals since Jan 1, 1601 UTC)
        def decode_filetime(ft_bytes):
            try:
                ft_int = int.from_bytes(ft_bytes, 'little')
                if ft_int == 0: return None
                
                EPOCH_AS_FILETIME = 116444736000000000
                HUNDREDS_OF_NANOSECONDS = 10000000
                
                if ft_int < EPOCH_AS_FILETIME: return None
                
                ts = (ft_int - EPOCH_AS_FILETIME) / HUNDREDS_OF_NANOSECONDS
                # Use local time for display consistent with user expectation
                dt = datetime.datetime.fromtimestamp(ts)
                return dt.strftime("%Y-%m-%d %H:%M:%S")
            except:
                return None

        # 1. Check End Time at 0x90
        if len(header_bytes) >= 0x98:
            d = decode_filetime(header_bytes[0x90:0x98])
            if d:
                found_date = d

        # 2. Check Start Time at 0x88 (if End Time not found)
        if not found_date and len(header_bytes) >= 0x90:
            d = decode_filetime(header_bytes[0x88:0x90])
            if d:
                found_date = d

        # 3. Legacy SYSTEMTIME scan
        if not found_date:
            def check_systemtime(buf, offset):
                if offset + 16 > len(buf): return None
                # SYSTEMTIME: 8 unsigned shorts (Year, Month, Dow, Day, Hour, Min, Sec, Ms)
                vals = struct.unpack('<8H', buf[offset:offset+16])
                wYear, wMonth, wDow, wDay, wHour, wMinute, wSecond, wMs = vals
                
                # Loose validation for reasonable date
                if (1990 <= wYear <= 2030 and 
                    1 <= wMonth <= 12 and 
                    1 <= wDay <= 31 and 
                    0 <= wHour <= 23 and 
                    0 <= wMinute <= 59 and 
                    0 <= wSecond <= 59):
                        # Filter out empty/zero dates
                        if wYear == 0 and wMonth == 0: return None
                        return f"{wYear}-{wMonth:02d}-{wDay:02d} {wHour:02d}:{wMinute:02d}:{wSecond:02d}"
                return None

            # Check known gaps
            for i in range(0xA0, 0xF0, 2):
                 d = check_systemtime(header_bytes, i)
                 if d:
                     found_date = d
                     break
            
            # Scan full header if needed
            if not found_date:
                for i in range(0, 512, 2):
                    d = check_systemtime(header_bytes, i)
                    if d:
                        found_date = d
                        break
                    
    except Exception as e:
        print(f"Error scanning WDF date: {e}")
        
    # 3. Fallback to file modification time
    if not found_date:
        try:
            fname = getattr(reader.file_obj, 'name', None)
            if fname and os.path.exists(fname):
                ts = os.path.getmtime(fname)
                dt = datetime.datetime.fromtimestamp(ts)
                found_date = dt.strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            pass
            
    if found_date:
        metadata['timestamp'] = found_date
    
    return metadata


def view_text(ui, title, text):
        """ Create a QTextBrowser to display a text content"""
        report_viewer = QDialog(ui)
        report_viewer.setWindowTitle(title)
        report_viewer.setGeometry(100, 100, 800, 600)
        text_browser = QTextBrowser(report_viewer)
        text_browser.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        text_browser.setOpenExternalLinks(True)
        text_browser.setPlainText(text)
        text_browser.moveCursor(QTextCursor.Start)
        layout = QVBoxLayout(report_viewer)
        layout.addWidget(text_browser)
        report_viewer.show()  

        
def rgba_to_default_color(rgba, default_colors=DEFAULT_COLORS):
    """
    Convert an RGBA tuple to the closest color in DEFAULT_COLORS.
    If no DEFAULT_COLORS are given, falls back to hex.
    """
    import matplotlib.colors as mcolors
    # Convert input to RGB array
    rgb = np.array(mcolors.to_rgb(rgba)) # drops alpha

    # Compute distance to each default color
    best_color = None
    best_dist = float("inf")
    for dc in default_colors:
        dc_rgb = np.array(mcolors.to_rgb(dc))
        dist = np.linalg.norm(rgb - dc_rgb)  # Euclidean distance in RGB
        if dist < best_dist:
            best_dist = dist
            best_color = dc

    return best_color if best_color else mcolors.to_hex(rgba)


def set_spectrum_item_color(item: QListWidgetItem, spectrum_info: dict):
    """Set list item background color based on spectrum status.
    
    Status hierarchy:
    1. Fitted (has_fit):
       - Converged: Color 4 (Soft Success Green)
       - Not Converged: Color 3 (Soft Warning Orange)
    2. Baselined (has_baseline): Color 2 (Soft Purple)
    3. Cropped (is_cropped): Color 1 (Distinct Gray)
    4. Original/Reinit: Transparent (no color)
    """
    is_cropped = spectrum_info.get("is_cropped", False)
    has_baseline = spectrum_info.get("has_baseline", False)
    has_fit = spectrum_info.get("has_fit", False)
    fit_success = spectrum_info.get("fit_success", False)

    if has_fit:
        if fit_success:
            # Color 4: Converged Fit (Distinct Green)
            item.setBackground(QColor(76, 175, 80, 120))
        else:
            # Color 3: Unconverged Fit (Distinct Red)
            item.setBackground(QColor(244, 67, 54, 120))
    elif has_baseline:
        # Color 2: Baselined (Distinct Purple)
        item.setBackground(QColor(156, 39, 176, 120))
    elif is_cropped:
        # Color 1: Cropped (Distinct Gray)
        item.setBackground(QColor(158, 158, 158, 120))
    else:
        # Original/Reinit: Transparent
        item.setBackground(QColor(0, 0, 0, 0))

def show_alert(message):
    """Show alert"""
    msg_box = QMessageBox()
    msg_box.setIcon(QMessageBox.Warning)
    msg_box.setWindowTitle("Alert")
    msg_box.setText(message)
    msg_box.exec_()


def show_toast_notification(parent, message, title=None, duration=3000, preset=None):
    """Show an auto-dismissing toast notification"""
    try:
        from pyqttoast import Toast, ToastPreset
    except ImportError:
        # Fallback to console print if pyqttoast not available
        prefix = f"[{title}] " if title else ""
        print(f"{prefix}{message}")
        return None
    
    toast = Toast(parent)
    toast.setDuration(duration)
    if title:
        toast.setTitle(title)
    toast.setText(message)
    
    # Apply preset style (default to SUCCESS)
    if preset is None:
        preset = ToastPreset.SUCCESS
    toast.applyPreset(preset)
    
    toast.show()
    return toast



def closest_index(array, value):
    return int(np.abs(array - value).argmin())


def parse_coords_from_fname(fname: str) -> tuple[float, float] | None:
    """Parse (x, y) spatial coordinates from a map spectrum fname.

    Spectra are named ``"{map_name}_({x}, {y})"`` (see
    ``VMWorkspaceMaps._extract_spectra_from_map``). Returns the coordinates as a
    float tuple, or None if the fname carries no parseable ``(x, y)`` suffix.
    """
    if '(' not in fname or ')' not in fname:
        return None
    coords_str = fname[fname.rfind('(') + 1:fname.rfind(')')]
    try:
        x_str, y_str = coords_str.split(',')
        return (float(x_str.strip()), float(y_str.strip()))
    except (ValueError, AttributeError):
        return None


def fano_display_amplitude(internal_ampli, q):
    """Convert a Fano peak's internal fit amplitude to its displayed peak height.

    The Fano lineshape's 'ampli' fit parameter scales the whole curve rather
    than the peak height directly (see fit_engine/models.py's batched_fano);
    the actual maximum intensity is ampli * (q**2 + 1). Used wherever the UI
    shows or edits a Fano peak's amplitude as an intuitive peak height
    instead of the raw fit parameter.
    """
    return internal_ampli * (q ** 2 + 1)


def fano_internal_amplitude(display_ampli, q):
    """Inverse of fano_display_amplitude(): displayed peak height -> internal fit amplitude."""
    return display_ampli / (q ** 2 + 1)


def save_df_to_excel(save_path, df):
    """Saves a DataFrame to an Excel file with colored columns based on prefixes."""
    if not save_path:
        return False, "No save path provided."

    try:
        import pandas as pd
        from openpyxl.styles import PatternFill
        if df.empty:
            return False, "DataFrame is empty. Nothing to save."
        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Results')
            workbook = writer.book
            worksheet = writer.sheets['Results']
            palette = ['#bda16d', '#a27ba0', '#cb5b12', '#23993b', '#008281', '#147ce4']
            prefix_colors = {}

            # Apply colors based on column prefixes
            for col_idx, col_name in enumerate(df.columns, start=1):
                prefix = col_name.split("_")[0] if "_" in col_name else col_name
                if prefix not in prefix_colors:
                    color_index = len(prefix_colors) % len(palette)
                    prefix_colors[prefix] = PatternFill(start_color=palette[color_index][1:], 
                                                        end_color=palette[color_index][1:], 
                                                        fill_type='solid')
                for row in range(2, len(df) + 2):  # Apply the fill color to the entire column
                    worksheet.cell(row=row, column=col_idx).fill = prefix_colors[prefix]

        return True, "DataFrame saved successfully."
    
    except Exception as e:
        return False, f"Error when saving DataFrame: {str(e)}"
    
def replace_peak_labels(fit_model, param):
        """Replace prefix 'm01' of peak model by labels designed by user"""
        peak_labels = fit_model["peak_labels"]
        if "_" in param:
            prefix, param = param.split("_", 1)  
            # Convert prefix to peak_label
            peak_index = int(prefix[1:]) - 1
            if 0 <= peak_index < len(peak_labels):
                peak_label = peak_labels[peak_index]
                return f"{param}_{peak_label}"
        return param

def copy_fig_to_clb(canvas, size_ratio=None):
    """Copy matplotlib figure canvas to clipboard with optional resizing. """
    if not canvas:
        QMessageBox.critical(None, "Error", "No plot to copy.")
        return
    
    figure = canvas.figure
    original_size = figure.get_size_inches()
    original_dpi = figure.dpi
    
    try:
        # Optionally resize the figure
        if size_ratio:
            figure.set_size_inches(size_ratio, forward=True)
            canvas.draw()
        buf = BytesIO()
        figure.savefig(buf, format='png', dpi=300, bbox_inches='tight')
        buf.seek(0)
        
        # Load PNG from buffer and copy to clipboard
        qimage = QImage()
        qimage.loadFromData(buf.getvalue())
        
        clipboard = QApplication.clipboard()
        clipboard.setPixmap(QPixmap.fromImage(qimage))
        
    finally:
        # Restore original figure state
        figure.set_size_inches(original_size, forward=True)
        figure.set_dpi(original_dpi)
        canvas.draw()

# Vector formats where embedding real (editable/searchable) text instead of
# outlined paths matters for publication use.
_VECTOR_FORMATS = ('pdf', 'eps', 'svg')

def export_figure_to_file(canvas, filepath, fmt, dpi=300, transparent=False, size_inches=None):
    """Export a matplotlib figure canvas to a file (PNG/TIFF/SVG/PDF/EPS).

    Mirrors copy_fig_to_clb's temporarily-resize-then-restore pattern so a
    caller can request a specific physical export size without permanently
    resizing the on-screen figure. Vector formats get font-embedding
    rcParams (fonttype 42 for PDF/EPS, real <text> elements for SVG) so
    exported text stays editable/searchable rather than outlined -- applied
    only for the duration of this save via matplotlib.rc_context, not
    process-wide.

    Returns True on success, False (after showing an error dialog) on failure.
    """
    if not canvas:
        QMessageBox.critical(None, "Error", "No plot to export.")
        return False

    figure = canvas.figure
    original_size = figure.get_size_inches()
    original_dpi = figure.dpi

    try:
        if size_inches:
            figure.set_size_inches(size_inches, forward=True)
            canvas.draw()

        # bbox_inches='tight' (copy_fig_to_clb's convenience default, avoids
        # excess whitespace) recomputes the saved bbox from rendered content
        # -- which silently ignores an explicitly requested physical size.
        # Only use it when the caller did NOT ask for a specific size; honor
        # the exact requested dimensions otherwise.
        bbox_inches = None if size_inches else 'tight'
        savefig_kwargs = dict(format=fmt, dpi=dpi, transparent=transparent, bbox_inches=bbox_inches)

        if fmt in _VECTOR_FORMATS:
            with matplotlib.rc_context({'pdf.fonttype': 42, 'ps.fonttype': 42, 'svg.fonttype': 'none'}):
                figure.savefig(filepath, **savefig_kwargs)
        else:
            figure.savefig(filepath, **savefig_kwargs)
        return True
    except Exception as e:
        QMessageBox.critical(None, "Error", f"Error exporting figure: {e}")
        return False
    finally:
        # Restore original figure state
        figure.set_size_inches(original_size, forward=True)
        figure.set_dpi(original_dpi)
        canvas.draw()

def build_clean_fit_model(fit_model):
    """Build a clean, consistently ordered fit model dict for serialization.

    Key order: fit_params, range_min, range_max, baseline, peak_labels, peak_models.
    Also rounds float values in fit_params to avoid QSettings float32 precision artifacts
    (e.g. 9.999999747378752e-06 -> 1e-05).
    """
    # Round fit_params floats to 6 significant digits to avoid float32 artifacts
    fit_params = fit_model.get("fit_params", {})
    if fit_params:
        cleaned_params = {}
        for k, v in fit_params.items():
            cleaned_params[k] = float(f"{v:.6g}") if isinstance(v, float) else v
        fit_params = cleaned_params

    # Clean baseline: strip fields not needed for the current mode
    baseline = fit_model.get("baseline")
    if baseline and isinstance(baseline, dict):
        baseline = deepcopy(baseline)
        baseline.pop("y_eval", None)
        if baseline.get("mode") == "Linear":
            baseline.pop("coef", None)
            baseline.pop("order_max", None)

    # Build ordered dict
    return {
        "fit_params": fit_params,
        "range_min": fit_model.get("range_min"),
        "range_max": fit_model.get("range_max"),
        "baseline": baseline,
        "peak_labels": fit_model.get("peak_labels", []),
        "peak_models": fit_model.get("peak_models", {}),
    }
